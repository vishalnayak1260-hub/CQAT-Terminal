import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import yfinance as yf
import requests
from fpdf import FPDF
import scipy.stats as si
import scipy.optimize as sco
from google import genai
import io
from datetime import datetime
import json
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# ==========================================
# 1. PAGE CONFIGURATION & CUSTOM CSS
# ==========================================
st.set_page_config(page_title="Titan Equity Terminal", page_icon="💼", layout="wide")

st.markdown("""
    <style>
    .main {background-color: #0e1117;}
    h1, h2, h3 {color: #e2e8f0; font-family: 'Helvetica Neue', sans-serif;}
    .stMetric {background-color: #1e293b; padding: 15px; border-radius: 8px; border: 1px solid #334155; box-shadow: 0 4px 6px rgba(0,0,0,0.1);}
    .stTabs [data-baseweb="tab"] {font-size: 13px; font-weight: 600; color: #94a3b8; padding: 12px 16px;}
    .stTabs [data-baseweb="tab"][aria-selected="true"] {color: #3b82f6; border-bottom-color: #3b82f6;}
    .command-bar {background-color: #1e293b; padding: 20px; border-radius: 10px; margin-bottom: 20px; border: 1px solid #3b82f6;}
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 2. CORE ENGINE & MEMORY CACHING
# ==========================================
def resolve_automatic_peer(ticker, sector="", industry=""):
    """Returns a single direct symmetrical peer for pairs trading (Tab 12)."""
    tk = ticker.upper().strip()
    pairs_map = {
        "TCS.NS": "INFY.NS", "INFY.NS": "TCS.NS", "WIPRO.NS": "TCS.NS", "HCLTECH.NS": "INFY.NS",
        "BIOCON.NS": "DRREDDY.NS", "DRREDDY.NS": "CIPLA.NS", "CIPLA.NS": "SUNPHARMA.NS", "SUNPHARMA.NS": "CIPLA.NS",
        "TATAMOTORS.NS": "M&M.NS", "M&M.NS": "TATAMOTORS.NS", "MARUTI.NS": "M&M.NS",
        "RELIANCE.NS": "ONGC.NS", "HDFCBANK.NS": "ICICIBANK.NS", "ICICIBANK.NS": "HDFCBANK.NS", "SBIN.NS": "BOB.NS",
        "AAPL": "MSFT", "MSFT": "AAPL", "GOOG": "META", "META": "GOOG", "AMZN": "WMT", "KO": "PEP", "PEP": "KO",
        "NVDA": "AMD", "AMD": "NVDA", "JPM": "BAC", "BAC": "JPM", "XOM": "CVX", "CVX": "XOM"
    }
    if tk in pairs_map: return pairs_map[tk]
    sec, ind = sector.lower(), industry.lower()
    is_indian = ".NS" in tk or ".BO" in tk
    if "software" in ind or "it services" in ind or "technology" in sec: return "INFY.NS" if is_indian else "MSFT"
    if "pharma" in ind or "biotech" in ind or "health" in sec: return "CIPLA.NS" if is_indian else "PFE"
    if "bank" in ind or "finance" in sec: return "ICICIBANK.NS" if is_indian else "JPM"
    if "auto" in ind or "vehicle" in ind: return "M&M.NS" if is_indian else "F"
    if "energy" in sec or "oil" in ind: return "ONGC.NS" if is_indian else "XOM"
    return "INFY.NS" if is_indian else "AAPL"

def resolve_peer_cohort(ticker, sector="", industry=""):
    """
    Returns an autofilled cohort of 3 to 5 relevant direct industry peers.
    Used to dynamically populate Comps (Tab 2) and Portfolio Optimization (Tab 6).
    """
    tk = ticker.upper().strip()
    is_indian = ".NS" in tk or ".BO" in tk

    # Specific Flagship Institutional Cohorts
    cohorts_map = {
        # US Tech & Mega-Cap
        "AAPL": ["MSFT", "GOOG", "AMZN", "META"],
        "MSFT": ["AAPL", "GOOG", "AMZN", "ORCL"],
        "GOOG": ["META", "MSFT", "AMZN", "AAPL"],
        "META": ["GOOG", "SNAP", "PINS", "MSFT"],
        "AMZN": ["WMT", "COST", "TGT", "BABA"],
        "NVDA": ["AMD", "INTC", "TSM", "QCOM", "AVGO"],
        "AMD": ["NVDA", "INTC", "QCOM", "AVGO"],
        "INTC": ["AMD", "NVDA", "TSM", "QCOM"],
        "TSLA": ["F", "GM", "TM", "RIVN"],
        # US Financials
        "JPM": ["BAC", "WFC", "C", "MS", "GS"],
        "BAC": ["JPM", "WFC", "C", "USB"],
        "V": ["MA", "AXP", "PYPL"],
        # US Healthcare & Pharma
        "PFE": ["JNJ", "MRK", "ABBV", "BMY"],
        "LLY": ["NVO", "AZN", "MRK", "PFE"],
        "UNH": ["ELV", "CVS", "CI", "HUM"],
        # US Energy & Industrials
        "XOM": ["CVX", "COP", "BP", "SHEL"],
        "CVX": ["XOM", "COP", "BP", "OXY"],
        "CAT": ["DE", "PCAR", "CMI"],
        # US Consumer
        "KO": ["PEP", "MNST", "KDP", "MDLZ"],
        "PEP": ["KO", "KDP", "MNST", "MDLZ"],
        "MCD": ["YUM", "SBUX", "QSR", "DPZ"],
        # Indian IT Services
        "TCS.NS": ["INFY.NS", "WIPRO.NS", "HCLTECH.NS", "TECHM.NS"],
        "INFY.NS": ["TCS.NS", "WIPRO.NS", "HCLTECH.NS", "LTIM.NS"],
        "WIPRO.NS": ["TCS.NS", "INFY.NS", "HCLTECH.NS", "TECHM.NS"],
        "HCLTECH.NS": ["TCS.NS", "INFY.NS", "WIPRO.NS", "LTIM.NS"],
        "TECHM.NS": ["TCS.NS", "INFY.NS", "WIPRO.NS", "HCLTECH.NS"],
        # Indian Banking & Finance
        "HDFCBANK.NS": ["ICICIBANK.NS", "SBIN.NS", "KOTAKBANK.NS", "AXISBANK.NS"],
        "ICICIBANK.NS": ["HDFCBANK.NS", "SBIN.NS", "AXISBANK.NS", "KOTAKBANK.NS"],
        "SBIN.NS": ["HDFCBANK.NS", "ICICIBANK.NS", "BOB.NS", "PNB.NS"],
        "KOTAKBANK.NS": ["HDFCBANK.NS", "ICICIBANK.NS", "AXISBANK.NS", "SBIN.NS"],
        "AXISBANK.NS": ["ICICIBANK.NS", "HDFCBANK.NS", "KOTAKBANK.NS", "SBIN.NS"],
        "BAJFINANCE.NS": ["BAJAJFINSV.NS", "HDFCBANK.NS", "CHOLAFIN.NS", "MUTHOOTFIN.NS"],
        # Indian Healthcare & Pharma
        "BIOCON.NS": ["DRREDDY.NS", "CIPLA.NS", "SUNPHARMA.NS", "LUPIN.NS"],
        "CIPLA.NS": ["SUNPHARMA.NS", "DRREDDY.NS", "DIVISLAB.NS", "TORNTPHARM.NS"],
        "DRREDDY.NS": ["CIPLA.NS", "SUNPHARMA.NS", "DIVISLAB.NS", "LUPIN.NS"],
        "SUNPHARMA.NS": ["DRREDDY.NS", "CIPLA.NS", "DIVISLAB.NS", "TORNTPHARM.NS"],
        # Indian Auto
        "TATAMOTORS.NS": ["M&M.NS", "MARUTI.NS", "BAJAJ-AUTO.NS", "HEROMOTOCO.NS"],
        "M&M.NS": ["TATAMOTORS.NS", "MARUTI.NS", "EICHERMOT.NS", "BAJAJ-AUTO.NS"],
        "MARUTI.NS": ["TATAMOTORS.NS", "M&M.NS", "HYUNDAI.NS", "BAJAJ-AUTO.NS"],
        # Indian Energy & Conglomerate
        "RELIANCE.NS": ["ONGC.NS", "BPCL.NS", "IOC.NS", "TCS.NS", "BHARTIARTL.NS"],
        "ONGC.NS": ["OIL.NS", "GAIL.NS", "BPCL.NS", "IOC.NS"],
        "ITC.NS": ["HINDUNILVR.NS", "NESTLEIND.NS", "BRITANNIA.NS", "DABUR.NS"],
        "HINDUNILVR.NS": ["ITC.NS", "NESTLEIND.NS", "BRITANNIA.NS", "GODREJCP.NS"],
        "BHARTIARTL.NS": ["TATACOMM.NS", "IDEA.NS", "RELIANCE.NS"]
    }
    if tk in cohorts_map: return cohorts_map[tk]

    # Sector & Industry Heuristics
    sec, ind = sector.lower(), industry.lower()
    if "software" in ind or "it services" in ind or "technology" in sec:
        return ["TCS.NS", "INFY.NS", "WIPRO.NS", "HCLTECH.NS"] if is_indian else ["MSFT", "ORCL", "CRM", "ADBE"]
    if "pharma" in ind or "biotech" in ind or "health" in sec:
        return ["SUNPHARMA.NS", "DRREDDY.NS", "CIPLA.NS", "DIVISLAB.NS"] if is_indian else ["PFE", "JNJ", "MRK", "ABBV"]
    if "bank" in ind or "finance" in sec:
        return ["HDFCBANK.NS", "ICICIBANK.NS", "SBIN.NS", "AXISBANK.NS"] if is_indian else ["JPM", "BAC", "WFC", "C"]
    if "auto" in ind or "vehicle" in ind:
        return ["TATAMOTORS.NS", "M&M.NS", "MARUTI.NS"] if is_indian else ["F", "GM", "TSLA", "TM"]
    if "energy" in sec or "oil" in ind:
        return ["ONGC.NS", "BPCL.NS", "IOC.NS"] if is_indian else ["XOM", "CVX", "COP", "SLB"]
    if "consumer" in sec or "food" in ind:
        return ["HINDUNILVR.NS", "ITC.NS", "NESTLEIND.NS"] if is_indian else ["PG", "KO", "PEP", "COST"]

    return ["TCS.NS", "HDFCBANK.NS", "RELIANCE.NS", "INFY.NS"] if is_indian else ["AAPL", "MSFT", "GOOG", "AMZN"]

def run_adf_stationarity_test(series, max_lag=1):
    """
    Augmented Dickey-Fuller stationarity test using pure NumPy OLS regression.
    Tests H0: unit root exists (non-stationary) vs H1: series is stationary / mean-reverting.
    Returns: (t_stat, approximate_p_value, is_cointegrated_bool)
    """
    y = np.asarray(series).astype(float)
    n = len(y)
    if n < 15:
        return 0.0, 0.50, False
    dy = np.diff(y)
    y_lag = y[max_lag:n-1]
    dy_curr = dy[max_lag:]
    
    X = [np.ones(len(y_lag)), y_lag]
    for i in range(1, max_lag + 1):
        X.append(dy[max_lag-i:len(dy)-i])
    X = np.column_stack(X)
    
    try:
        coeffs, _, _, _ = np.linalg.lstsq(X, dy_curr, rcond=None)
        gamma = coeffs[1]
        resids = dy_curr - X @ coeffs
        dof = len(dy_curr) - X.shape[1]
        if dof <= 0: return 0.0, 0.50, False
        sigma2 = np.sum(resids**2) / dof
        var_coeffs = sigma2 * np.linalg.inv(X.T @ X)
        se_gamma = np.sqrt(max(var_coeffs[1, 1], 1e-12))
        t_stat = gamma / se_gamma
        
        # MacKinnon 1%, 5%, 10% critical threshold anchors for constant (no trend)
        crit_1, crit_5, crit_10 = -3.43, -2.86, -2.57
        if t_stat < crit_1:
            p_val = 0.01
        elif t_stat < crit_5:
            p_val = 0.05 - (crit_5 - t_stat) / (crit_5 - crit_1) * 0.04
        elif t_stat < crit_10:
            p_val = 0.10 - (crit_10 - t_stat) / (crit_10 - crit_5) * 0.05
        else:
            p_val = min(1.0, 0.10 + (t_stat - crit_10) * 0.15)
            
        is_cointegrated = (t_stat < crit_5)
        return float(t_stat), float(max(p_val, 0.001)), bool(is_cointegrated)
    except Exception:
        return 0.0, 0.50, False

@st.cache_data(ttl=300, show_spinner=False)
def search_asset_candidates(query):
    clean_query = query.strip()
    if not clean_query: return []
    url = f"https://query2.finance.yahoo.com/v1/finance/search?q={clean_query}"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    candidates = []
    try:
        response = requests.get(url, headers=headers, timeout=5)
        data = response.json()
        if 'quotes' in data:
            for quote in data['quotes']:
                if quote.get('quoteType') in ['EQUITY', 'ETF', 'INDEX']:
                    name = quote.get('shortname', quote.get('longname', 'Unknown Name'))
                    symbol = quote.get('symbol', '')
                    exch = quote.get('exchange', 'Unknown Exchange')
                    candidates.append({"display": f"{name} ({symbol}) - {exch}", "symbol": symbol})
    except Exception: pass
    
    if not candidates:
        raw = clean_query.upper().replace(" ", "")
        candidates.append({"display": f"Direct Ticker Match: {raw}", "symbol": raw})
        if "." not in raw:
            candidates.append({"display": f"Indian Market Guess: {raw}.NS", "symbol": f"{raw}.NS"})
    return candidates

@st.cache_data(ttl=3600, show_spinner=False)
def pull_market_action(ticker_symbol, time_horizon="1y"):
    try:
        df = yf.download(ticker_symbol, period=time_horizon, progress=False)
        if df.empty: return False, None
        
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
            
        df['SMA_50'] = df['Close'].rolling(window=50).mean()
        df['SMA_200'] = df['Close'].rolling(window=200).mean()
        df = df.dropna().reset_index()
        
        date_col = 'Date' if 'Date' in df.columns else 'Datetime'
        df_market = pd.DataFrame({
            'date': pd.to_datetime(df[date_col]).dt.tz_localize(None),
            'close_price': df['Close'].astype(float),
            'volume': df['Volume'].astype(int),
            'sma_50': df['SMA_50'].astype(float),
            'sma_200': df['SMA_200'].astype(float)
        })
        return True, df_market
    except Exception:
        return False, None

@st.cache_data(ttl=3600, show_spinner=False)
def pull_peer_action(tickers_list, time_horizon="1y"):
    try:
        df = yf.download(tickers_list, period=time_horizon, progress=False)
        if df.empty: return pd.DataFrame()
        
        if isinstance(df.columns, pd.MultiIndex):
            if 'Close' in df.columns.get_level_values(0):
                return df['Close']
            elif 'Close' in df.columns.get_level_values(1):
                return df.xs('Close', level=1, axis=1)
        elif 'Close' in df.columns:
            res = df[['Close']].copy()
            res.columns = [tickers_list[0]]
            return res
        return pd.DataFrame()
    except Exception: return pd.DataFrame()

@st.cache_data(ttl=3600, show_spinner=False)
def pull_macro_regime(time_horizon="1y"):
    try:
        df = yf.download(['^TNX', '^IRX', '^VIX'], period=time_horizon, progress=False)
        if 'Close' in df.columns: return df['Close']
        return pd.DataFrame()
    except Exception: return pd.DataFrame()

@st.cache_data(ttl=3600, show_spinner=False)
def pull_institutional_profile(ticker):
    """Delegates to yfinance's native engine to utilize curl_cffi cloud bypass."""
    info_dict = {}
    tk = yf.Ticker(ticker)
    
    try:
        fast = tk.fast_info
        info_dict['marketCap'] = getattr(fast, 'market_cap', None)
        info_dict['sharesOutstanding'] = getattr(fast, 'shares', None)
        info_dict['fiftyTwoWeekHigh'] = getattr(fast, 'year_high', None)
        info_dict['fiftyTwoWeekLow'] = getattr(fast, 'year_low', None)
        info_dict['previousClose'] = getattr(fast, 'previous_close', None)
        info_dict['currency'] = getattr(fast, 'currency', None)
        info_dict['exchange'] = getattr(fast, 'exchange', None)
    except Exception: pass

    try:
        fetched = tk.info
        if isinstance(fetched, dict) and 'symbol' in fetched:
            for k, v in fetched.items():
                if v is not None and (k not in info_dict or info_dict[k] is None):
                    info_dict[k] = v
    except Exception: pass
        
    return info_dict

def extract_financial_statements(raw_ticker, info):
    """
    Extracts core financial statements, capturing actual Interest Expense,
    Tax Provision, Pre-Tax Income, Cash, and Debt for exact fundamental modeling.
    """
    metrics = {
        'revenue': info.get('totalRevenue'), 'net_income': info.get('netIncomeToCommon'), 
        'total_assets': info.get('totalAssets'), 'total_equity': info.get('totalStockholderEquity'), 
        'fcf': info.get('freeCashflow'), 'total_liabilities': info.get('totalLiabilitiesNetMinorityInterest'),
        'ebit': info.get('ebitda'), 'operating_cashflow': info.get('operatingCashflow'),
        'current_assets': None, 'current_liabilities': None,
        'interest_expense': None, 'tax_provision': None, 'pretax_income': None,
        'cash_and_equivalents': info.get('totalCash'), 'total_debt': info.get('totalDebt')
    }
    try:
        inc, bs, cf = raw_ticker.financials, raw_ticker.balance_sheet, raw_ticker.cashflow
        
        # Revenue & Net Income
        if not inc.empty and metrics['revenue'] is None:
            for k in ['Total Revenue', 'Operating Revenue', 'Revenue']:
                if k in inc.index: metrics['revenue'] = inc.loc[k].iloc[0]; break
        if not inc.empty and metrics['net_income'] is None:
            for k in ['Net Income', 'Net Income Common Stockholders']:
                if k in inc.index: metrics['net_income'] = inc.loc[k].iloc[0]; break
                
        # Interest Expense, Tax Provision, Pretax Income (for real Kd and real Tax Rate)
        if not inc.empty:
            for k in ['Interest Expense', 'Interest Expense Non Operating', 'Total Interest Expense']:
                if k in inc.index: metrics['interest_expense'] = abs(inc.loc[k].iloc[0]); break
            for k in ['Tax Provision', 'Income Tax Expense', 'Provision for Income Taxes']:
                if k in inc.index: metrics['tax_provision'] = abs(inc.loc[k].iloc[0]); break
            for k in ['Pretax Income', 'Income Before Tax', 'Earnings Before Tax']:
                if k in inc.index: metrics['pretax_income'] = inc.loc[k].iloc[0]; break
                
        # Balance Sheet line items
        if not bs.empty and metrics['total_assets'] is None:
            if 'Total Assets' in bs.index: metrics['total_assets'] = bs.loc['Total Assets'].iloc[0]
        if not bs.empty and metrics['total_equity'] is None:
            for k in ['Stockholders Equity', 'Common Stock Equity']:
                if k in bs.index: metrics['total_equity'] = bs.loc[k].iloc[0]; break
        if not bs.empty and metrics['total_liabilities'] is None:
            if 'Total Liabilities Net Minority Interest' in bs.index: metrics['total_liabilities'] = bs.loc['Total Liabilities Net Minority Interest'].iloc[0]
        if not bs.empty and metrics['current_assets'] is None:
            if 'Total Current Assets' in bs.index: metrics['current_assets'] = bs.loc['Total Current Assets'].iloc[0]
        if not bs.empty and metrics['current_liabilities'] is None:
            if 'Total Current Liabilities' in bs.index: metrics['current_liabilities'] = bs.loc['Total Current Liabilities'].iloc[0]
        if not bs.empty and metrics['cash_and_equivalents'] is None:
            for k in ['Cash And Cash Equivalents', 'Cash Cash Equivalents And Short Term Investments']:
                if k in bs.index: metrics['cash_and_equivalents'] = bs.loc[k].iloc[0]; break
        if not bs.empty and metrics['total_debt'] is None:
            for k in ['Total Debt', 'Long Term Debt And Capital Lease Obligation']:
                if k in bs.index: metrics['total_debt'] = bs.loc[k].iloc[0]; break

        # Cash Flow line items
        if not cf.empty and metrics['fcf'] is None:
            if 'Free Cash Flow' in cf.index: metrics['fcf'] = cf.loc['Free Cash Flow'].iloc[0]
        if not cf.empty and metrics['operating_cashflow'] is None:
            if 'Operating Cash Flow' in cf.index: metrics['operating_cashflow'] = cf.loc['Operating Cash Flow'].iloc[0]
    except Exception: pass
    return metrics

def extract_multi_year_financials(raw_ticker):
    """Pulls up to 4 fiscal years of core line items for the Excel model historical financials tab."""
    years, hist = [], {
        'revenue': [], 'net_income': [], 'ebit': [], 'ocf': [], 'fcf': [],
        'total_assets': [], 'total_equity': [], 'total_liabilities': [],
        'current_assets': [], 'current_liabilities': []
    }
    try:
        inc, bs, cf = raw_ticker.financials, raw_ticker.balance_sheet, raw_ticker.cashflow
        if inc.empty and bs.empty:
            return years, hist
        cols = inc.columns if not inc.empty else bs.columns
        years = [c.strftime('%Y') if hasattr(c, 'strftime') else str(c) for c in cols][:4]
        n = len(years)

        def pull(df, keys):
            row = None
            for k in keys:
                if not df.empty and k in df.index:
                    row = df.loc[k]
                    break
            if row is None:
                return [None] * n
            vals = list(row.values)[:n]
            return vals + [None] * (n - len(vals))

        hist['revenue'] = pull(inc, ['Total Revenue', 'Operating Revenue', 'Revenue'])
        hist['net_income'] = pull(inc, ['Net Income', 'Net Income Common Stockholders'])
        hist['ebit'] = pull(inc, ['EBITDA', 'Normalized EBITDA'])
        hist['ocf'] = pull(cf, ['Operating Cash Flow'])
        hist['fcf'] = pull(cf, ['Free Cash Flow'])
        hist['total_assets'] = pull(bs, ['Total Assets'])
        hist['total_equity'] = pull(bs, ['Stockholders Equity', 'Common Stock Equity'])
        hist['total_liabilities'] = pull(bs, ['Total Liabilities Net Minority Interest'])
        hist['current_assets'] = pull(bs, ['Total Current Assets', 'Current Assets'])
        hist['current_liabilities'] = pull(bs, ['Total Current Liabilities', 'Current Liabilities'])
    except Exception:
        pass
    return years, hist

# Palette & Fonts for Excel
NAVY = "1E293B"
BLUE_HDR = "1D4ED8"
LIGHT_BLUE = "DBEAFE"
YELLOW = "FFF9C4"
GREY = "F1F5F9"
WHITE = "FFFFFF"
INPUT_BLUE = "0000FF"
LINK_GREEN = "007A33"
BLACK = "000000"

FONT_NAME = "Calibri"
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name=FONT_NAME, size=10, color=WHITE)
SECTION_FONT = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True, color=NAVY)
NOTE_FONT = Font(name=FONT_NAME, size=8, italic=True, color="64748B")
INPUT_FONT = Font(name=FONT_NAME, size=10, color=INPUT_BLUE, bold=False)
FORMULA_FONT = Font(name=FONT_NAME, size=10, color=BLACK)
LINK_FONT = Font(name=FONT_NAME, size=10, color=LINK_GREEN)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=BLUE_HDR)
YELLOW_FILL = PatternFill("solid", fgColor=YELLOW)
GREY_FILL = PatternFill("solid", fgColor=GREY)

THIN = Side(style="thin", color="CBD5E1")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR_FMT = '#,##0.00;(#,##0.00);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'
MULT_FMT = '0.00"x"'
NUM_FMT = '#,##0;(#,##0);"-"'

def _banner(ws, text, subtitle=None, span=8, row=1, height=32):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height
    if subtitle:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=span)
        c2 = ws.cell(row=row + 1, column=1, value=subtitle)
        c2.font = SUBTITLE_FONT
        c2.fill = HEADER_FILL
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row + 1].height = 18

def _section(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    return row + 1

def _kv(ws, row, label, value, col=1, fmt=None, font=None, fill=None, note=None):
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = LABEL_FONT
    vc = ws.cell(row=row, column=col + 1, value=value)
    vc.font = font or FORMULA_FONT
    if fmt: vc.number_format = fmt
    if fill: vc.fill = fill
    vc.border = BOX
    if note:
        nc = ws.cell(row=row, column=col + 2, value=note)
        nc.font = NOTE_FONT
    return row + 1

def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _page_setup(ws):
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

def generate_excel_model(ctx: dict) -> bytes:
    wb = Workbook()
    ticker = ctx.get("ticker", "N/A")
    full_name = ctx.get("full_name", ticker)
    curr_sym = ctx.get("curr_sym", "$")
    currency = ctx.get("currency", "USD")
    today_str = datetime.now().strftime("%d %b %Y")

    # ============================================================ SUMMARY ==
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [28, 16, 16, 18, 15, 15, 15, 15])
    _page_setup(ws)
    _banner(ws, f"{full_name} ({ticker}) — Institutional Financial Model",
            f"Sector: {ctx.get('sector','N/A')}  |  Industry: {ctx.get('industry','N/A')}  |  "
            f"Exchange: {ctx.get('exchange','N/A')}  |  Generated {today_str}", span=8)

    r = 4
    r = _section(ws, r, "Market Snapshot", span=8)
    price_row = r
    r = _kv(ws, r, "Current Price", ctx.get("current_price", 0), fmt=CUR_FMT, font=INPUT_FONT)
    r = _kv(ws, r, "Market Cap ($mm)", (ctx.get("market_cap") or 0) / 1_000_000, fmt=NUM_FMT, font=INPUT_FONT)
    r = _kv(ws, r, "52-Week High", ctx.get("high_52", 0), fmt=CUR_FMT, font=INPUT_FONT)
    r = _kv(ws, r, "52-Week Low", ctx.get("low_52", 0), fmt=CUR_FMT, font=INPUT_FONT)
    r = _kv(ws, r, "Trailing P/E", ctx.get("pe", 0), fmt=MULT_FMT, font=INPUT_FONT)
    r = _kv(ws, r, "Beta", ctx.get("beta", 1.0), fmt='0.00', font=INPUT_FONT)
    r = _kv(ws, r, "Dividend Yield", ctx.get("div_yield", 0), fmt=PCT_FMT, font=INPUT_FONT)
    r = _kv(ws, r, "Shares Outstanding (mm)", (ctx.get("shares_out") or 0) / 1_000_000, fmt=NUM_FMT, font=INPUT_FONT)
    r += 1

    r = _section(ws, r, "Valuation Summary (linked to DCF Model)", span=8)
    headers = ["Scenario", "Implied Price", "Current Price", "Upside / (Downside)"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=1 + i, value=h)
        c.font = LABEL_FONT; c.fill = GREY_FILL; c.border = BOX
    r += 1
    scenarios = ["Bear Case", "Base Case", "Bull Case"]
    summary_dcf_rows = {}
    for sc in scenarios:
        summary_dcf_rows[sc] = r
        ws.cell(row=r, column=1, value=sc).font = FORMULA_FONT
        c2 = ws.cell(row=r, column=2, value=0); c2.font = LINK_FONT; c2.number_format = CUR_FMT
        c3 = ws.cell(row=r, column=3, value=f"=$B${price_row}"); c3.font = FORMULA_FONT; c3.number_format = CUR_FMT
        c4 = ws.cell(row=r, column=4, value=f"=(B{r}-C{r})/C{r}"); c4.font = FORMULA_FONT; c4.number_format = PCT_FMT
        for cc in range(1, 5): ws.cell(row=r, column=cc).border = BOX
        r += 1
    base_case_row = summary_dcf_rows["Base Case"]
    r += 1

    r = _section(ws, r, "Signal Check", span=8)
    ws.cell(row=r, column=1, value="Graham Number (Defensive Value)").font = LABEL_FONT
    gc = ws.cell(row=r, column=2, value=ctx.get("graham_number", 0)); gc.font = INPUT_FONT; gc.number_format = CUR_FMT
    r += 1
    ws.cell(row=r, column=1, value="Verdict vs. Base Case DCF").font = LABEL_FONT
    vc = ws.cell(row=r, column=2, value=f'=IF(B{base_case_row}>C{base_case_row},"Undervalued","Overvalued")')
    vc.font = FORMULA_FONT
    r += 2
    note = ws.cell(row=r, column=1,
                    value="Blue = source input   Black = formula   Green = cross-sheet link   "
                          "Yellow = key assumption to edit. Edit the Assumptions tab to flex the model.")
    note.font = NOTE_FONT

    # ========================================================= ASSUMPTIONS ==
    ws2 = wb.create_sheet("Assumptions")
    ws2.sheet_view.showGridLines = False
    _set_widths(ws2, [30, 16, 4, 40])
    _page_setup(ws2)
    _banner(ws2, "Assumptions & Cost of Capital", "Edit the yellow cells — every model tab recalculates.", span=4)

    r = 4
    r = _section(ws2, r, "Capital Structure & Discount Rate", span=4)
    a_rf = r; _kv(ws2, r, "Risk-Free Rate", ctx.get("rf", 0.045), fmt=PCT_FMT, font=INPUT_FONT, fill=YELLOW_FILL); r += 1
    a_beta = r; _kv(ws2, r, "Beta", ctx.get("beta", 1.0), fmt='0.00', font=INPUT_FONT, fill=YELLOW_FILL); r += 1
    a_erp = r; _kv(ws2, r, "Equity Risk Premium", ctx.get("erp", 0.055), fmt=PCT_FMT, font=INPUT_FONT, fill=YELLOW_FILL); r += 1
    a_ke = r; ws2.cell(row=r, column=1, value="Cost of Equity (Ke)").font = LABEL_FONT
    ce = ws2.cell(row=r, column=2, value=f"=B{a_rf}+(B{a_beta}*B{a_erp})"); ce.font = FORMULA_FONT; ce.number_format = PCT_FMT; ce.border = BOX
    r += 1
    a_kd = r; _kv(ws2, r, "Pre-Tax Cost of Debt (Kd)", ctx.get("cost_of_debt", 0.065), fmt=PCT_FMT, font=INPUT_FONT, fill=YELLOW_FILL, note="Calculated from Interest Exp / Total Debt"); r += 1
    a_tax = r; _kv(ws2, r, "Effective Tax Rate", ctx.get("tax_rate", 0.25), fmt=PCT_FMT, font=INPUT_FONT, fill=YELLOW_FILL, note="Calculated from Tax Provision / Pretax Income"); r += 1
    a_debt = r; _kv(ws2, r, "Total Debt ($mm)", (ctx.get("total_debt") or 0) / 1_000_000, fmt=NUM_FMT, font=INPUT_FONT); r += 1
    a_mcap = r; _kv(ws2, r, "Market Cap ($mm)", (ctx.get("market_cap") or 0) / 1_000_000, fmt=NUM_FMT, font=INPUT_FONT); r += 1
    a_we = r; ws2.cell(row=r, column=1, value="Weight of Equity").font = LABEL_FONT
    we = ws2.cell(row=r, column=2, value=f"=B{a_mcap}/(B{a_mcap}+B{a_debt})"); we.font = FORMULA_FONT; we.number_format = PCT_FMT; we.border = BOX
    r += 1
    a_wd = r; ws2.cell(row=r, column=1, value="Weight of Debt").font = LABEL_FONT
    wd = ws2.cell(row=r, column=2, value=f"=1-B{a_we}"); wd.font = FORMULA_FONT; wd.number_format = PCT_FMT; wd.border = BOX
    r += 1
    a_wacc = r; ws2.cell(row=r, column=1, value="WACC").font = LABEL_FONT
    wacc_c = ws2.cell(row=r, column=2, value=f"=(B{a_we}*B{a_ke})+(B{a_wd}*B{a_kd}*(1-B{a_tax}))")
    wacc_c.font = FORMULA_FONT; wacc_c.number_format = PCT_FMT; wacc_c.border = BOX; wacc_c.fill = PatternFill("solid", fgColor="D1FAE5")
    r += 2

    r = _section(ws2, r, "DCF Growth Scenarios", span=4)
    a_fcf = r; _kv(ws2, r, "Base FCF ($mm)", ctx.get("fcf_base", 0), fmt=NUM_FMT, font=INPUT_FONT, fill=YELLOW_FILL); r += 1
    a_bear = r; _kv(ws2, r, "Bear Case Growth (5yr)", ctx.get("growth_bear", 0.04), fmt=PCT_FMT, font=INPUT_FONT, fill=YELLOW_FILL); r += 1
    a_base = r; _kv(ws2, r, "Base Case Growth (5yr)", ctx.get("growth_base", 0.08), fmt=PCT_FMT, font=INPUT_FONT, fill=YELLOW_FILL); r += 1
    a_bull = r; _kv(ws2, r, "Bull Case Growth (5yr)", ctx.get("growth_bull", 0.14), fmt=PCT_FMT, font=INPUT_FONT, fill=YELLOW_FILL); r += 1
    a_term = r; _kv(ws2, r, "Terminal Growth Rate", ctx.get("terminal_growth", 0.04), fmt=PCT_FMT, font=INPUT_FONT, fill=YELLOW_FILL); r += 1
    a_shares = r; _kv(ws2, r, "Shares Outstanding (mm)", (ctx.get("shares_out") or 0) / 1_000_000, fmt=NUM_FMT, font=INPUT_FONT); r += 2

    A = "Assumptions"

    # ============================================================ DCF MODEL ==
    ws3 = wb.create_sheet("DCF Model")
    ws3.sheet_view.showGridLines = False
    _set_widths(ws3, [26, 14, 14, 14, 14, 14, 14])
    _page_setup(ws3)
    _banner(ws3, "Discounted Cash Flow — 5-Year FCF Projection", "Three-scenario build, linked to Assumptions tab.", span=7)

    r = 4
    r = _section(ws3, r, "Year", span=7)
    ws3.cell(row=r, column=1, value="Scenario").font = LABEL_FONT
    for y in range(1, 6):
        c = ws3.cell(row=r, column=1 + y, value=f"Year {y}")
        c.font = LABEL_FONT; c.fill = GREY_FILL; c.border = BOX
    r += 1

    fcf_rows = {}
    for sc, growth_row in [("Bear Case", a_bear), ("Base Case", a_base), ("Bull Case", a_bull)]:
        ws3.cell(row=r, column=1, value=f"{sc} — Projected FCF ($mm)").font = FORMULA_FONT
        for y in range(1, 6):
            col = 1 + y
            col_l = get_column_letter(col)
            prev_l = get_column_letter(col - 1)
            formula = f"='{A}'!$B${a_fcf}*(1+'{A}'!$B${growth_row})" if y == 1 else f"={prev_l}{r}*(1+'{A}'!$B${growth_row})"
            cc = ws3.cell(row=r, column=col, value=formula)
            cc.font = LINK_FONT; cc.number_format = NUM_FMT; cc.border = BOX
        fcf_rows[sc] = r
        r += 1
    r += 1

    r = _section(ws3, r, "Present Value & Implied Share Price", span=7)
    for i, h in enumerate(["Scenario", "PV of 5-Yr FCF", "Terminal Value", "PV of TV", "Enterprise Value", "Implied Price"]):
        c = ws3.cell(row=r, column=1 + i, value=h); c.font = LABEL_FONT; c.fill = GREY_FILL; c.border = BOX
    r += 1
    dcf_summary_rows = {}
    for sc in ["Bear Case", "Base Case", "Bull Case"]:
        fr = fcf_rows[sc]
        ws3.cell(row=r, column=1, value=sc).font = FORMULA_FONT
        pv_terms = "+".join([f"{get_column_letter(1+y)}{fr}/(1+'{A}'!$B${a_wacc})^{y}" for y in range(1, 6)])
        pv_c = ws3.cell(row=r, column=2, value=f"={pv_terms}"); pv_c.font = LINK_FONT; pv_c.number_format = NUM_FMT; pv_c.border = BOX
        y5 = f"{get_column_letter(6)}{fr}"
        tv_c = ws3.cell(row=r, column=3, value=f"={y5}*(1+'{A}'!$B${a_term})/('{A}'!$B${a_wacc}-'{A}'!$B${a_term})")
        tv_c.font = LINK_FONT; tv_c.number_format = NUM_FMT; tv_c.border = BOX
        pvtv_c = ws3.cell(row=r, column=4, value=f"=C{r}/(1+'{A}'!$B${a_wacc})^5"); pvtv_c.font = FORMULA_FONT; pvtv_c.number_format = NUM_FMT; pvtv_c.border = BOX
        ev_c = ws3.cell(row=r, column=5, value=f"=B{r}+D{r}"); ev_c.font = FORMULA_FONT; ev_c.number_format = NUM_FMT; ev_c.border = BOX
        ip_c = ws3.cell(row=r, column=6, value=f"=E{r}/'{A}'!$B${a_shares}"); ip_c.font = FORMULA_FONT; ip_c.number_format = CUR_FMT; ip_c.border = BOX
        ip_c.fill = PatternFill("solid", fgColor="D1FAE5")
        dcf_summary_rows[sc] = r
        r += 1

    for sc, target_row in summary_dcf_rows.items():
        target_cell = ws.cell(row=target_row, column=2)
        target_cell.value = f"='DCF Model'!F{dcf_summary_rows[sc]}"
        target_cell.font = LINK_FONT; target_cell.number_format = CUR_FMT

    # ================================================== HISTORICAL FINANCIALS ==
    ws4 = wb.create_sheet("Historical Financials")
    ws4.sheet_view.showGridLines = False
    hist_years = ctx.get("hist_years", ["FY"])
    hist = ctx.get("hist", {})
    n_years = len(hist_years)
    _set_widths(ws4, [26] + [14] * n_years)
    _page_setup(ws4)
    _banner(ws4, "Historical Financial Statements ($mm)", "Source: Company filings via Yahoo Finance.", span=1 + n_years)

    r = 4
    r = _section(ws4, r, "Income Statement & Cash Flow", span=1 + n_years)
    ws4.cell(row=r, column=1, value="Line Item ($mm)").font = LABEL_FONT; ws4.cell(row=r, column=1).fill = GREY_FILL; ws4.cell(row=r, column=1).border = BOX
    for i, y in enumerate(hist_years):
        c = ws4.cell(row=r, column=2 + i, value=str(y)); c.font = LABEL_FONT; c.fill = GREY_FILL; c.border = BOX; c.alignment = Alignment(horizontal="center")
    r += 1
    line_items = [
        ("revenue", "Total Revenue"), ("net_income", "Net Income"), ("ebit", "EBITDA"),
        ("ocf", "Operating Cash Flow"), ("fcf", "Free Cash Flow"),
    ]
    row_ref = {}
    for key, label in line_items:
        ws4.cell(row=r, column=1, value=label).font = FORMULA_FONT
        vals = hist.get(key, [None] * n_years)
        for i in range(n_years):
            v = vals[i] if i < len(vals) and vals[i] is not None else 0
            c = ws4.cell(row=r, column=2 + i, value=round(v / 1_000_000, 2) if v else 0)
            c.font = INPUT_FONT; c.number_format = NUM_FMT; c.border = BOX
        row_ref[key] = r
        r += 1

    # ================================================================ COMPS ==
    ws5 = wb.create_sheet("Comps")
    ws5.sheet_view.showGridLines = False
    comps = ctx.get("comps", [])
    _set_widths(ws5, [34, 22, 14, 14, 16])
    _page_setup(ws5)
    _banner(ws5, "Relative Valuation — Peer Comparables", "Source: Yahoo Finance, dynamic industry cohort.", span=5)
    r = 4
    r = _section(ws5, r, "Trading Comps", span=5)
    for i, h in enumerate(["Ticker", "Company", "P/E", "EV/EBITDA", "Net Margin"]):
        c = ws5.cell(row=r, column=1 + i, value=h); c.font = LABEL_FONT; c.fill = GREY_FILL; c.border = BOX
    r += 1
    comp_start = r
    for row_data in comps:
        ws5.cell(row=r, column=1, value=row_data.get("ticker", "")).font = INPUT_FONT
        ws5.cell(row=r, column=2, value=row_data.get("name", "")).font = INPUT_FONT
        pe_c = ws5.cell(row=r, column=3, value=row_data.get("pe") or None); pe_c.font = INPUT_FONT; pe_c.number_format = MULT_FMT
        ev_c = ws5.cell(row=r, column=4, value=row_data.get("ev_ebitda") or None); ev_c.font = INPUT_FONT; ev_c.number_format = MULT_FMT
        nm_c = ws5.cell(row=r, column=5, value=(row_data.get("net_margin") or 0) / 100 if row_data.get("net_margin") else None)
        nm_c.font = INPUT_FONT; nm_c.number_format = PCT_FMT
        for cc in range(1, 6): ws5.cell(row=r, column=cc).border = BOX
        r += 1
    comp_end = r - 1
    if comp_end >= comp_start:
        ws5.cell(row=r, column=2, value="Peer Average").font = LABEL_FONT
        for col, fmt in [(3, MULT_FMT), (4, MULT_FMT), (5, PCT_FMT)]:
            col_l = get_column_letter(col)
            c = ws5.cell(row=r, column=col, value=f"=IFERROR(AVERAGE({col_l}{comp_start}:{col_l}{comp_end}),0)")
            c.font = FORMULA_FONT; c.number_format = fmt; c.border = BOX
        r += 1
        ws5.cell(row=r, column=2, value="Peer Median").font = LABEL_FONT
        for col, fmt in [(3, MULT_FMT), (4, MULT_FMT), (5, PCT_FMT)]:
            col_l = get_column_letter(col)
            c = ws5.cell(row=r, column=col, value=f"=IFERROR(MEDIAN({col_l}{comp_start}:{col_l}{comp_end}),0)")
            c.font = FORMULA_FONT; c.number_format = fmt; c.border = BOX

    # ============================================================= HEALTH ==
    ws6 = wb.create_sheet("Health & Ratios")
    ws6.sheet_view.showGridLines = False
    _set_widths(ws6, [30, 16, 4, 40])
    _page_setup(ws6)
    _banner(ws6, "Corporate Health — DuPont, Altman Z, Piotroski", span=4)
    r = 4
    r = _section(ws6, r, "DuPont ROE Decomposition", span=4)
    dupont = ctx.get("dupont", {})
    if dupont.get("valid"):
        r = _kv(ws6, r, "Net Profit Margin", dupont.get("npm", 0) / 100, fmt=PCT_FMT, font=INPUT_FONT)
        r = _kv(ws6, r, "Asset Turnover", dupont.get("ato", 0), fmt=MULT_FMT, font=INPUT_FONT)
        r = _kv(ws6, r, "Equity Multiplier", dupont.get("em", 0), fmt=MULT_FMT, font=INPUT_FONT)
        dp_row = r
        ws6.cell(row=r, column=1, value="ROE (NPM × ATO × EM)").font = LABEL_FONT
        c = ws6.cell(row=r, column=2, value=f"=B{dp_row-3}*B{dp_row-2}*B{dp_row-1}")
        c.font = FORMULA_FONT; c.number_format = PCT_FMT; c.border = BOX; c.fill = PatternFill("solid", fgColor="D1FAE5")
        r += 2

    # ======================================================= PRICE HISTORY ==
    ws7 = wb.create_sheet("Price History")
    ws7.sheet_view.showGridLines = False
    price_df = ctx.get("price_df")
    _set_widths(ws7, [14, 14, 14, 14, 14])
    _page_setup(ws7)
    _banner(ws7, f"Historical Price Action — {ticker}", "Raw daily series with 50 / 200-day moving averages.", span=5)
    r = 4
    headers = ["Date", "Close", "SMA 50", "SMA 200", "Volume"]
    for i, h in enumerate(headers):
        c = ws7.cell(row=r, column=1 + i, value=h); c.font = LABEL_FONT; c.fill = GREY_FILL; c.border = BOX
    data_start = r + 1
    if price_df is not None and not price_df.empty:
        for _, row_data in price_df.iterrows():
            r += 1
            ws7.cell(row=r, column=1, value=row_data["date"].strftime("%Y-%m-%d") if hasattr(row_data["date"], "strftime") else str(row_data["date"])).font = INPUT_FONT
            ws7.cell(row=r, column=2, value=round(float(row_data["close_price"]), 2)).font = INPUT_FONT; ws7.cell(row=r, column=2).number_format = CUR_FMT
            ws7.cell(row=r, column=3, value=round(float(row_data["sma_50"]), 2) if pd_notna(row_data.get("sma_50")) else None).number_format = CUR_FMT
            ws7.cell(row=r, column=4, value=round(float(row_data["sma_200"]), 2) if pd_notna(row_data.get("sma_200")) else None).number_format = CUR_FMT
            ws7.cell(row=r, column=5, value=int(row_data["volume"])).font = INPUT_FONT; ws7.cell(row=r, column=5).number_format = NUM_FMT
        data_end = r

        chart = LineChart()
        chart.title = f"{ticker} — Price & Moving Averages"
        chart.style = 2
        chart.y_axis.title = f"Price ({currency})"
        chart.x_axis.title = "Date"
        chart.height = 9; chart.width = 22
        data_ref = Reference(ws7, min_col=2, max_col=4, min_row=4, max_row=data_end)
        cats_ref = Reference(ws7, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws7.add_chart(chart, f"G4")

    ws.sheet_view.tabSelected = True
    wb.active = 0

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def pd_notna(v):
    try:
        import math
        if v is None: return False
        return not (isinstance(v, float) and math.isnan(v))
    except Exception: return v is not None

def _safe(s): return str(s).encode("latin-1", "ignore").decode("latin-1")

def _mini_chart(df_market, accent_hex="#3b82f6"):
    fig, ax = plt.subplots(figsize=(7.6, 2.3), dpi=200)
    ax.plot(df_market["date"], df_market["close_price"], color=accent_hex, linewidth=1.4, label="Close")
    if "sma_50" in df_market.columns:
        ax.plot(df_market["date"], df_market["sma_50"], color="#ef4444", linewidth=0.9, label="SMA 50")
    if "sma_200" in df_market.columns:
        ax.plot(df_market["date"], df_market["sma_200"], color="#10b981", linewidth=0.9, label="SMA 200")
    ax.legend(loc="upper left", fontsize=6, frameon=False)
    ax.tick_params(labelsize=6, colors="#475569")
    for spine in ["top", "right"]: ax.spines[spine].set_visible(False)
    ax.grid(axis="y", linewidth=0.3, alpha=0.4)
    fig.tight_layout(pad=0.4)
    buf = io.BytesIO()
    fig.savefig(buf, format="png", transparent=True)
    plt.close(fig)
    buf.seek(0)
    return buf

class TearSheet(FPDF):
    def footer(self):
        self.set_y(-14)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(100, 116, 139)
        self.cell(0, 5, "Prepared using Titan Equity Terminal. Independently generated from public market data; not investment advice.", align="C")
        self.set_y(-9)
        self.cell(0, 5, f"Generated {datetime.now().strftime('%d %b %Y, %H:%M')} | Page {self.page_no()}", align="C")

def generate_pdf_report(ctx: dict) -> bytes:
    curr_sym = ctx.get("curr_sym", "$")
    pdf_sym = "INR " if curr_sym == "₹" else curr_sym
    ticker = ctx.get("ticker", "N/A")
    full_name = _safe(ctx.get("full_name", ticker))
    sector = _safe(ctx.get("sector", "N/A"))
    industry = _safe(ctx.get("industry", "N/A"))
    exchange = _safe(ctx.get("exchange", "N/A"))
    c_price = ctx.get("current_price", 0) or 0
    dcf_results = ctx.get("dcf_results", {}) or {}
    graham = ctx.get("graham_number", 0) or 0

    pdf = TearSheet(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.add_page()
    pdf.set_margins(12, 12, 12)

    # Header
    pdf.set_fill_color(30, 41, 59)
    pdf.rect(0, 0, 210, 28, "F")
    pdf.set_xy(12, 7); pdf.set_text_color(255, 255, 255); pdf.set_font("Helvetica", "B", 16)
    pdf.cell(140, 8, f"{full_name} ({ticker})", ln=0)
    pdf.set_font("Helvetica", "", 9); pdf.set_xy(12, 16)
    pdf.cell(140, 6, f"{sector}  |  {industry}  |  {exchange}", ln=0)

    # Rating badge
    base_case = dcf_results.get("Base Case")
    if base_case and c_price:
        upside = (base_case - c_price) / c_price
        badge, bg = ("UNDERVALUED", (16, 185, 129)) if upside > 0.10 else (("OVERVALUED", (239, 68, 68)) if upside < -0.10 else ("FAIRLY VALUED", (245, 158, 11)))
    else:
        badge, bg = "N/A", (100, 116, 139)
    pdf.set_fill_color(*bg); pdf.rect(155, 8, 43, 12, "F"); pdf.set_xy(155, 8)
    pdf.set_font("Helvetica", "B", 11); pdf.set_text_color(255, 255, 255)
    pdf.cell(43, 12, badge, align="C")

    pdf.set_text_color(0, 0, 0); pdf.set_y(33)

    # Snapshot strip
    stats = [
        ("Price", f"{pdf_sym}{c_price:,.2f}"),
        ("Market Cap", f"{pdf_sym}{(ctx.get('market_cap') or 0)/1e9:,.2f}B"),
        ("Beta", f"{ctx.get('beta', 1.0):.2f}"),
        ("P/E (TTM)", f"{ctx.get('pe', 0) or 0:.1f}x" if ctx.get("pe") else "N/A"),
        ("52W Range", f"{pdf_sym}{ctx.get('low_52', 0):,.0f} - {pdf_sym}{ctx.get('high_52', 0):,.0f}"),
        ("Div Yield", f"{(ctx.get('div_yield') or 0)*100:.2f}%"),
    ]
    col_w = 186 / len(stats)
    pdf.set_fill_color(241, 245, 249); pdf.rect(12, pdf.get_y(), 186, 16, "F")
    y0 = pdf.get_y()
    for i, (label, val) in enumerate(stats):
        x = 12 + i * col_w
        pdf.set_xy(x, y0 + 2); pdf.set_font("Helvetica", "", 7); pdf.set_text_color(100, 116, 139)
        pdf.cell(col_w, 4, label, align="C")
        pdf.set_xy(x, y0 + 7); pdf.set_font("Helvetica", "B", 10); pdf.set_text_color(30, 41, 59)
        pdf.cell(col_w, 6, _safe(val), align="C")
    pdf.set_text_color(0, 0, 0); pdf.set_y(y0 + 20)

    # Chart
    df_market = ctx.get("df_market")
    if df_market is not None and not df_market.empty:
        try:
            chart_buf = _mini_chart(df_market)
            pdf.image(chart_buf, x=12, y=pdf.get_y(), w=186)
            pdf.set_y(pdf.get_y() + 58)
        except Exception: pass

    # DCF & Fundamentals
    pdf.set_fill_color(59, 130, 246); pdf.set_text_color(255, 255, 255); pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, "  Intrinsic Valuation (DCF Architecture)", ln=True, fill=True)
    pdf.set_text_color(0, 0, 0); pdf.ln(1)
    if dcf_results:
        widths = [46, 46, 46, 48]
        pdf.set_font("Helvetica", "B", 8); pdf.set_fill_color(241, 245, 249)
        for w, lab in zip(widths, ["Scenario", "Implied Price", "vs. Current", "Upside / (Downside)"]):
            pdf.cell(w, 6, lab, border=0, fill=True)
        pdf.ln(); pdf.set_font("Helvetica", "", 9)
        for scenario, price in dcf_results.items():
            upside = ((price - c_price) / c_price) * 100 if c_price > 0 else 0
            pdf.cell(widths[0], 6, _safe(scenario))
            pdf.cell(widths[1], 6, f"{pdf_sym}{price:,.2f}")
            pdf.cell(widths[2], 6, f"{pdf_sym}{c_price:,.2f}")
            pdf.set_text_color(*( (16, 185, 129) if upside >= 0 else (239, 68, 68) ))
            pdf.cell(widths[3], 6, f"{upside:+.1f}%")
            pdf.set_text_color(0, 0, 0); pdf.ln()
    pdf.ln(2)

    return pdf.output(dest="S").encode("latin-1")

def generate_ai_thesis(ticker, full_name, metrics_summary):
    client = genai.Client(api_key=st.secrets["GEMINI_API_KEY"])
    prompt = f"""You are a Lead Portfolio Manager at an elite institutional fund using a GARP (Growth at a Reasonable Price) framework.
Analyze these metrics for {full_name} ({ticker}):

{metrics_summary}

Analytical Directives:
1. DO NOT default to a 'SELL' just because the DCF value is lower than the current price. Great companies (wide moats, high ROIC, monopolies) deserve to trade at premium valuations.
2. Issue a 'BUY' if the company has a strong Piotroski F-Score proxy (3+ out of the 4 tracked signals: positive net income, positive OCF, OCF exceeding net income, positive ROA), a positive ROIC vs WACC spread, and is a dominant market player—even if the DCF implies a slight premium. 
3. Issue a 'HOLD' if the stock is highly overvalued but the underlying business is exceptional, or if it is fairly priced with average fundamentals.
4. Issue a 'SELL' ONLY if the underlying business is destroying capital (negative ROIC spread, poor F-Score, severe structural risks) AND it is overvalued.

Return STRICT, valid JSON adhering to this schema:
{{
    "verdict": "BUY", "HOLD", or "SELL",
    "target_rationale": "One concise, data-grounded sentence summarizing the core investment stance.",
    "variant_perception": "Key operational or market dynamic that the consensus narrative is mispricing.",
    "valuation_case": "Concise synthesis of intrinsic DCF scenario targets vs current market trading price and WACC.",
    "core_risk_factor": "The single most critical structural, financial, or competitive risk identified in the data.",
    "quantitative_grounding": ["List 2-3 specific metrics that proved your point"]
}}"""
    try:
        response = client.models.generate_content(
            model="gemini-3.5-flash", contents=prompt, config={"response_mime_type": "application/json"}
        )
    except Exception as e:
        raise RuntimeError(f"Gemini API call failed: {e}") from e

    candidates = getattr(response, "candidates", None)
    if not candidates: raise RuntimeError("Gemini returned no candidates — prompt may have been blocked.")
    finish_reason = str(getattr(candidates[0], "finish_reason", "") or "")
    if finish_reason and "STOP" not in finish_reason.upper():
        raise RuntimeError(f"Gemini stopped early (finish_reason={finish_reason}).")

    raw_text = (response.text or "").strip()
    if not raw_text: raise RuntimeError("Gemini returned an empty response body.")

    cleaned = raw_text
    if cleaned.startswith("```"):
        cleaned = cleaned.strip("`").strip()
        if cleaned.lower().startswith("json"): cleaned = cleaned[4:].strip()

    try:
        return json.loads(cleaned)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"Gemini response was not valid JSON ({e}). Raw: {cleaned[:120]!r}") from e

@st.cache_data(ttl=1800, show_spinner=False)
def generate_ai_thesis_cached(ticker, full_name, metrics_summary):
    return generate_ai_thesis(ticker, full_name, metrics_summary)

# ==========================================
# 3. SIDEBAR: MACRO CONTROL DECK & DYNAMIC PEER AUTOFILTER
# ==========================================
st.sidebar.title("🎛️ Engine Config")
st.sidebar.markdown("Modify core institutional variables below.")

with st.sidebar.expander("🌍 Macro & Horizon Inputs", expanded=True):
    time_horizon = st.selectbox("Analysis Time Horizon:", ["1mo", "3mo", "6mo", "1y", "2y", "5y", "max"], index=3)
    benchmark_input = st.text_input("Macro Benchmark Overlay:", "^GSPC", help="Used to map relative momentum.")
    global_rf = st.number_input("US Risk-Free Rate (%)", value=4.5, step=0.1) / 100
    global_erp = st.number_input("Equity Risk Premium (%)", value=5.5, step=0.1) / 100

# ==========================================
# 4. MAIN BODY: COMMAND LINE SEARCH
# ==========================================
st.title("💼 Titan Institutional Research Terminal")

st.markdown('<div class="command-bar">', unsafe_allow_html=True)
cmd_col1, cmd_col2, cmd_col3 = st.columns([3, 1, 1])
with cmd_col1:
    raw_input = st.text_input("🔍 Command Line Search:", placeholder="Enter Keyword or Ticker (e.g., Tata, AAPL, TCS.NS)...", label_visibility="collapsed")
with cmd_col2:
    st.write("") 
    search_button = st.button("🚀 Execute Terminal", use_container_width=True)
st.markdown('</div>', unsafe_allow_html=True)

selected_ticker = None
if raw_input:
    candidates = search_asset_candidates(raw_input)
    if candidates:
        options_dict = {c["display"]: c["symbol"] for c in candidates}
        with cmd_col3:
            st.write("") 
            selection = st.selectbox("Verify Asset:", list(options_dict.keys()), label_visibility="collapsed")
            selected_ticker = options_dict[selection]
    else:
        st.warning("No public market matches found.")

# Dynamic Peer Cohort Synchronization
if selected_ticker:
    if "last_tracked_ticker" not in st.session_state:
        st.session_state.last_tracked_ticker = None
    
    # When a new ticker is selected, autofill industry-matched peers
    if selected_ticker != st.session_state.last_tracked_ticker:
        st.session_state.last_tracked_ticker = selected_ticker
        temp_prof = pull_institutional_profile(selected_ticker)
        auto_peers = resolve_peer_cohort(selected_ticker, temp_prof.get('sector', ''), temp_prof.get('industry', ''))
        st.session_state.peer_input_val = ", ".join(auto_peers)

# Sidebar Peer Input (Reflected with dynamic autofill)
with st.sidebar.expander("⚖️ Portfolio & Comps Targets", expanded=True):
    portfolio_capital = st.number_input("Initial Fund Capital", min_value=1000, value=1000000, step=10000)
    current_default_peers = st.session_state.get("peer_input_val", "MSFT, GOOG, AMZN, META")
    peer_input = st.text_input("Peer Tickers (Autofilled by Industry):", value=current_default_peers, 
                               help="Automatically populated with direct industry peers. Modify or add custom tickers as desired.")
    if st.button("🔄 Reset to Industry Peers", use_container_width=True):
        if selected_ticker:
            temp_prof = pull_institutional_profile(selected_ticker)
            st.session_state.peer_input_val = ", ".join(resolve_peer_cohort(selected_ticker, temp_prof.get('sector', ''), temp_prof.get('industry', '')))
            st.rerun()

if "app_running" not in st.session_state: st.session_state.app_running = False
if search_button and selected_ticker: st.session_state.app_running = True

# ==========================================
# 5. EXECUTION MATRIX & MODULES
# ==========================================
if st.session_state.app_running and selected_ticker:
    st.divider()
    with st.spinner(f"Connecting to Exchange and running Quantitative Engines for {selected_ticker}..."):
        is_success, df_market = pull_market_action(selected_ticker, time_horizon)
    
    if is_success and df_market is not None:
        raw_ticker = yf.Ticker(selected_ticker)
        info = pull_institutional_profile(selected_ticker).copy()
        full_name = info.get('longName', info.get('shortName', selected_ticker))
        
        deep_metrics = {
            'revenue': None, 'net_income': None, 'total_assets': None, 'total_equity': None, 
            'fcf': None, 'total_liabilities': None, 'ebit': None, 'operating_cashflow': None, 
            'current_assets': None, 'current_liabilities': None, 'interest_expense': None,
            'tax_provision': None, 'pretax_income': None, 'cash_and_equivalents': None, 'total_debt': None
        }
        try:
            extracted_metrics = extract_financial_statements(raw_ticker, info)
            deep_metrics.update(extracted_metrics)
        except: pass

        sector_str = str(info.get('sector', '')).lower()
        ind_str = str(info.get('industry', '')).lower()
        is_financial = 'bank' in ind_str or 'financial' in sector_str or 'insurance' in ind_str
        is_manufacturing = 'manufacturing' in sector_str or 'automotive' in ind_str or 'industrial' in sector_str
        is_indian = selected_ticker.endswith('.NS') or selected_ticker.endswith('.BO')

        currency = info.get('currency')
        if not currency:
            currency = 'INR' if is_indian else 'USD'
        info['currency'] = currency
        curr_sym = "₹" if currency == "INR" else "$"

        current_price = df_market['close_price'].iloc[-1] if not df_market.empty else info.get('currentPrice', 1.0)
        
        if not info.get('fiftyTwoWeekHigh') and not df_market.empty:
            recent_252 = df_market.tail(252)
            info['fiftyTwoWeekHigh'] = float(recent_252['close_price'].max())
            info['fiftyTwoWeekLow'] = float(recent_252['close_price'].min())
            
        shares = float(info.get('sharesOutstanding', 0) or 0)
        if not info.get('marketCap') and shares > 0 and current_price > 0:
            info['marketCap'] = current_price * shares

        beta_raw = info.get('beta', 1.0)
        raw_rev = float(deep_metrics.get('revenue', 1000000000.0) or 1000000000.0)
        raw_fcf = deep_metrics.get('fcf')
        fcf_base = float(raw_fcf) / 1000000.0 if raw_fcf is not None else (raw_rev * 0.12) / 1000000.0
        shares_for_calc = shares / 1000000.0 if shares > 0 else 100.0 
        
        eps, bvps = info.get('trailingEps', 0), info.get('bookValue', 0)
        graham_number = np.sqrt(22.5 * eps * bvps) if eps > 0 and bvps > 0 else 0.0
        dividend_yield = info.get('dividendYield', 0)
        dividend_rate = info.get('dividendRate', 0)
        
        # ------------------------------------------------------------------
        # FUNDAMENTAL PROXY REFINEMENTS (Real Kd, Real Tax Rate, True Invested Capital)
        # ------------------------------------------------------------------
        current_rf = 0.068 if is_indian else global_rf
        cost_of_equity = current_rf + (beta_raw * global_erp)
        
        # 1. Real Effective Cost of Debt (Kd)
        total_debt_val = float(deep_metrics.get('total_debt') or info.get('totalDebt') or 0.0)
        interest_exp = float(deep_metrics.get('interest_expense') or 0.0)
        if total_debt_val > 0 and interest_exp > 0:
            raw_kd = interest_exp / total_debt_val
            cost_of_debt = min(max(raw_kd, current_rf), 0.16) # Clamped within realistic corporate bounds
            kd_source = "Actual (Interest Exp / Total Debt)"
        else:
            cost_of_debt = current_rf + 0.02
            kd_source = "Sovereign Proxy (Rf + 2.0%)"

        # 2. Real Effective Corporate Tax Rate
        tax_prov = float(deep_metrics.get('tax_provision') or 0.0)
        pretax_inc = float(deep_metrics.get('pretax_income') or 0.0)
        if pretax_inc > 0 and tax_prov > 0:
            raw_tax = tax_prov / pretax_inc
            effective_tax_rate = min(max(raw_tax, 0.15), 0.35)
            tax_source = "Actual (Tax Provision / Pretax Income)"
        else:
            effective_tax_rate = 0.25 if is_indian else 0.21
            tax_source = "Statutory Corporate Proxy"

        # 3. True Operating Invested Capital
        total_equity_val = float(deep_metrics.get('total_equity') or 0.0)
        total_cash_val = float(deep_metrics.get('cash_and_equivalents') or info.get('totalCash') or 0.0)
        ta_val = float(deep_metrics.get('total_assets') or 0.0)
        tl_val = float(deep_metrics.get('total_liabilities') or 0.0)
        
        if total_equity_val > 0:
            invested_capital = total_equity_val + total_debt_val - total_cash_val
            ic_source = "Standard Operating (Equity + Debt - Cash)"
        elif ta_val > 0:
            invested_capital = ta_val - (tl_val * 0.4)
            ic_source = "Asset Proxy (Total Assets - 0.4*Liabilities)"
        else:
            invested_capital = 0.0
            ic_source = "N/A"

        # 4. WACC Calculation
        market_cap = float(info.get('marketCap') or (current_price * shares) or 0.0)
        if total_debt_val > 0 and market_cap > 0:
            total_capital = total_debt_val + market_cap
            weight_equity = market_cap / total_capital
            weight_debt = total_debt_val / total_capital
            calculated_wacc = (weight_equity * cost_of_equity) + (weight_debt * cost_of_debt * (1 - effective_tax_rate))
        else:
            calculated_wacc = cost_of_equity
            weight_equity, weight_debt = 1.0, 0.0

        # DCF Baseline FCF
        if raw_fcf is not None and float(raw_fcf) > 0:
            fcf_raw_base = float(raw_fcf)
        else:
            fcf_raw_base = raw_rev * 0.15  # 15% operating FCF margin proxy

        # DCF Scenario Builder with Enterprise-to-Equity Bridge
        pdf_dcf = {}
        if not is_financial:
            for n, g in {"Bear Case": 0.04, "Base Case": 0.08, "Bull Case": 0.14}.items():
                discounted_cfs = [
                    (fcf_raw_base * ((1 + g) ** y)) / ((1 + calculated_wacc) ** y)
                    for y in range(1, 6)
                ]
                pv_fcf = sum(discounted_cfs)
                terminal_fcf = fcf_raw_base * ((1 + g) ** 5) * (1 + 0.035)
                terminal_denom = max(calculated_wacc - 0.035, 0.02)
                tv = terminal_fcf / terminal_denom
                pv_tv = tv / ((1 + calculated_wacc) ** 5)
                enterprise_val = pv_fcf + pv_tv
                equity_val = enterprise_val + total_cash_val - total_debt_val
                safe_shares = max(shares, 1.0)
                pdf_dcf[n] = max(equity_val / safe_shares, 0.01)

        # DuPont Deconstruction
        ni, ta, te, rev = deep_metrics['net_income'], deep_metrics['total_assets'], deep_metrics['total_equity'], deep_metrics['revenue']
        dupont_data = {'valid': False}
        if ni and ta and te and rev and ta > 0 and te > 0 and rev > 0:
            dupont_data = {'valid': True, 'npm': (ni/rev)*100, 'ato': rev/ta, 'em': ta/te, 'roe': (ni/rev)*(rev/ta)*(ta/te)*100}

        # ROIC & Value Creation Spread
        ebit_val = float(deep_metrics.get('ebit') or 0.0)
        roic, roic_wacc_spread = None, None
        if ebit_val > 0 and invested_capital > 0:
            nopat = ebit_val * (1 - effective_tax_rate)
            roic = nopat / invested_capital
            roic_wacc_spread = roic - calculated_wacc

        # Altman Z-Score
        z_score, model_type, z_safe_limit, z_distress_limit = None, None, None, None
        if not is_financial:
            ca_h = deep_metrics.get('current_assets')
            cl_h = deep_metrics.get('current_liabilities')
            working_capital_h = (ca_h - cl_h) if ca_h and cl_h else 0
            if ta and rev and tl_val and market_cap and ebit_val and ta > 0 and tl_val > 0:
                x1 = working_capital_h / ta
                x2 = te / ta if te else 0.5
                x3 = ebit_val / ta
                x4 = market_cap / tl_val
                if is_manufacturing:
                    x5 = rev / ta
                    z_score = (1.2 * x1) + (1.4 * x2) + (3.3 * x3) + (0.6 * x4) + (1.0 * x5)
                    model_type = "Classic Manufacturing Z-Score"
                    z_safe_limit, z_distress_limit = 3.0, 1.8
                else:
                    z_score = (6.56 * x1) + (3.26 * x2) + (6.72 * x3) + (1.05 * x4)
                    model_type = "Emerging Service & Tech Z''-Score"
                    z_safe_limit, z_distress_limit = 2.6, 1.1

        # Piotroski F-Score Proxy
        f_score = 0
        if ni and ni > 0: f_score += 1
        if deep_metrics.get('operating_cashflow') and deep_metrics.get('operating_cashflow') > 0: f_score += 1
        if deep_metrics.get('operating_cashflow') and ni and deep_metrics.get('operating_cashflow') > ni: f_score += 1
        if info.get('returnOnAssets') and info.get('returnOnAssets') > 0: f_score += 1

        hist_years, hist_financials = extract_multi_year_financials(raw_ticker)

        # Header Display
        col_head1, col_head2 = st.columns([3, 1])
        with col_head1:
            st.header(f"📊 {full_name} ({selected_ticker})")
            st.markdown(f"**Sector:** {info.get('sector', 'N/A') or 'N/A'} | **Industry:** {info.get('industry', 'N/A') or 'N/A'} | **Exchange:** {info.get('exchange', 'N/A') or 'N/A'}")
        with col_head2:
            st.write("")
            if "ai_thesis_store" not in st.session_state: st.session_state.ai_thesis_store = {}
            pdf_slot = st.container()

            excel_ctx = {
                "ticker": selected_ticker, "full_name": full_name, "sector": info.get('sector', 'N/A'),
                "industry": info.get('industry', 'N/A'), "exchange": info.get('exchange', 'N/A'),
                "currency": currency, "curr_sym": curr_sym, "current_price": current_price,
                "market_cap": market_cap, "beta": beta_raw, "pe": info.get('trailingPE'),
                "div_yield": dividend_yield, "high_52": info.get('fiftyTwoWeekHigh'),
                "low_52": info.get('fiftyTwoWeekLow'), "shares_out": shares,
                "rf": current_rf, "erp": global_erp, "cost_of_debt": cost_of_debt,
                "tax_rate": effective_tax_rate, "total_debt": total_debt_val,
                "fcf_base": fcf_base, "wacc": calculated_wacc, "cost_of_equity": cost_of_equity,
                "growth_bear": 0.04, "growth_base": 0.08, "growth_bull": 0.14, "terminal_growth": 0.04,
                "graham_number": graham_number, "dividend_rate": dividend_rate,
                "hist_years": hist_years if hist_years else ["Latest"],
                "hist": hist_financials if hist_years else {k: [deep_metrics.get(k)] for k in
                    ['revenue', 'net_income', 'ebit', 'ocf', 'fcf', 'total_assets', 'total_equity',
                     'total_liabilities', 'current_assets', 'current_liabilities']},
                "dupont": dupont_data, "z_score": z_score, "z_model_type": model_type,
                "z_safe_limit": z_safe_limit, "z_distress_limit": z_distress_limit,
                "f_score": f_score, "roic": roic, "eps": eps,
                "comps": [], "price_df": df_market,
            }
            excel_bytes = generate_excel_model(excel_ctx)
            st.download_button(label="📥 Financial Model (XLSX)", data=excel_bytes, file_name=f"{selected_ticker}_Financial_Model.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        tabs = st.tabs([
            "📈 Market", "📊 Comps", "💎 Value", "🔮 Valuation", "🏛️ Health", "⚖️ Portfolio", 
            "🤖 Forecast", "🧮 Options", "📊 Algo", "📅 Season", "🛡️ VaR", "🎭 Arb", 
            "🕵️ Insiders", "🌐 Macro", "🧠 AI Thesis"
        ])
        tab_market, tab_comps, tab_value, tab_dcf, tab_health, tab_mpt, tab_ml, tab_bs, tab_tech, tab_season, tab_risk, tab_arb, tab_insider, tab_macro, tab_ai = tabs
        
        with tab_market:
            if not df_market.empty:
                m_col1, m_col2, m_col3, m_col4 = st.columns(4)
                m_col1.metric("Latest Close Price", f"{curr_sym}{current_price:.2f}")
                m_col2.metric("Trading Volume", f"{df_market['volume'].iloc[-1]:,}")
                m_col3.metric("Total Market Cap", f"{curr_sym}{market_cap / 1000000000:.2f} B" if market_cap else "N/A")
                m_col4.metric("Systematic Risk (Beta)", f"{info.get('beta', 1.0):.2f}")
                
                sm_col1, sm_col2, sm_col3, sm_col4 = st.columns(4)
                short_ratio = info.get('shortRatio', 'N/A')
                sm_col1.metric("Short Ratio (Days to Cover)", short_ratio if short_ratio else "N/A")
                
                high_52, low_52 = info.get('fiftyTwoWeekHigh'), info.get('fiftyTwoWeekLow')
                dist_to_high = ((current_price - high_52) / high_52) * 100 if high_52 and high_52 > 0 else 0
                dist_to_low = ((current_price - low_52) / low_52) * 100 if low_52 and low_52 > 0 else 0
                
                sm_col2.metric("52-Week High", f"{curr_sym}{float(high_52):.2f}" if high_52 else "N/A", f"{dist_to_high:.1f}%" if high_52 else "")
                sm_col3.metric("52-Week Low", f"{curr_sym}{float(low_52):.2f}" if low_52 else "N/A", f"{dist_to_low:.1f}%" if low_52 else "")
                
                inst_own = info.get('heldPercentInstitutions')
                sm_col4.metric("Institutional Ownership", f"{inst_own * 100:.2f}%" if inst_own else "N/A")
                
                st.markdown(f"#### Raw Price Action & Moving Averages ({selected_ticker})")
                fig_raw = px.line(df_market, x='date', y=['close_price', 'sma_50', 'sma_200'], color_discrete_sequence=['#3b82f6', '#ef4444', '#10b981'])
                fig_raw.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', xaxis_title="Date", yaxis_title=f"Price ({currency})", legend_title_text="Metrics")
                st.plotly_chart(fig_raw, use_container_width=True)

                st.markdown(f"#### Relative Macro Performance: {selected_ticker} vs {benchmark_input}")
                benchmark_df = pd.DataFrame()
                if benchmark_input:
                    try:
                        b_data = pull_peer_action([benchmark_input], time_horizon)
                        if not b_data.empty:
                            benchmark_df = pd.DataFrame(b_data).dropna().reset_index()
                            benchmark_df.columns = ['date', 'benchmark_close']
                            benchmark_df['date'] = pd.to_datetime(benchmark_df['date']).dt.tz_localize(None)
                    except: pass

                fig_price = go.Figure()
                start_price = df_market['close_price'].iloc[0]
                fig_price.add_trace(go.Scatter(x=df_market['date'], y=(df_market['close_price']/start_price)*100, name=selected_ticker, line=dict(color='#3b82f6', width=2)))
                if not benchmark_df.empty:
                    merged_df = pd.merge(df_market[['date']], benchmark_df, on='date', how='inner')
                    if not merged_df.empty:
                        start_bench = merged_df['benchmark_close'].iloc[0]
                        fig_price.add_trace(go.Scatter(x=merged_df['date'], y=(merged_df['benchmark_close']/start_bench)*100, name=f"Benchmark ({benchmark_input})", line=dict(color='#94a3b8', dash='dash')))
                fig_price.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', xaxis_title="Date", yaxis_title="Normalized Return %")
                st.plotly_chart(fig_price, use_container_width=True)

        with tab_comps:
            st.subheader("Relative Valuation Matrix")
            resolved_peers = []
            with st.spinner("Resolving industry peer cohort..."):
                for raw_p in [p.strip() for p in peer_input.split(",") if p.strip()]:
                    p_cands = search_asset_candidates(raw_p)
                    if p_cands and p_cands[0]['symbol'] not in resolved_peers and p_cands[0]['symbol'] != selected_ticker:
                        resolved_peers.append(p_cands[0]['symbol'])
            all_tickers_to_compare = [selected_ticker] + resolved_peers
            
            comps_data = []
            for t in all_tickers_to_compare:
                try:
                    p_info = pull_institutional_profile(t)
                    comps_data.append({"Ticker": t, "Company Name": p_info.get('shortName', t), 
                                       "P/E Ratio": p_info.get('trailingPE', None), 
                                       "EV/EBITDA": p_info.get('enterpriseToEbitda', None), 
                                       "Net Margin (%)": p_info.get('profitMargins', 0) * 100 if p_info.get('profitMargins') else None})
                except: pass
            
            if comps_data:
                df_comps = pd.DataFrame(comps_data)
                
                # Contextual Valuation Cards (Target Multiple vs Peer Benchmarks)
                target_pe = df_comps.loc[df_comps['Ticker'] == selected_ticker, 'P/E Ratio'].values[0]
                peer_pes = df_comps.loc[df_comps['Ticker'] != selected_ticker, 'P/E Ratio'].dropna()
                if len(peer_pes) > 0 and pd_notna(target_pe):
                    peer_avg_pe = peer_pes.mean()
                    pe_diff = ((target_pe - peer_avg_pe) / peer_avg_pe) * 100
                    c_col1, c_col2, c_col3 = st.columns(3)
                    c_col1.metric("Target P/E Ratio", f"{target_pe:.2f}x")
                    c_col2.metric("Industry Peer Average P/E", f"{peer_avg_pe:.2f}x")
                    if pe_diff > 0:
                        c_col3.metric("Valuation Spread vs. Peers", f"+{pe_diff:.1f}% Premium", delta=f"{pe_diff:.1f}%", delta_color="inverse")
                    else:
                        c_col3.metric("Valuation Spread vs. Peers", f"{pe_diff:.1f}% Discount", delta=f"{pe_diff:.1f}%", delta_color="normal")
                
                clean_scatter = df_comps.dropna(subset=['P/E Ratio', 'Net Margin (%)', 'EV/EBITDA']).copy()
                clean_scatter = clean_scatter[clean_scatter['EV/EBITDA'] > 0]
                
                if not clean_scatter.empty:
                    fig_scatter = px.scatter(clean_scatter, x="P/E Ratio", y="Net Margin (%)", text="Ticker", 
                                             size="EV/EBITDA", color="Ticker", title="Peer Positioning Scatter (Bubble Size = EV/EBITDA)")
                    fig_scatter.update_traces(textposition='top center')
                    fig_scatter.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig_scatter, use_container_width=True)
                
                df_comps_display = df_comps.fillna(np.nan)
                st.dataframe(
                    df_comps_display.style.highlight_max(axis=0, subset=['Net Margin (%)'])
                    .format({"P/E Ratio": "{:.2f}x", "EV/EBITDA": "{:.2f}x", "Net Margin (%)": "{:.2f}%"}, na_rep="N/A"), 
                    use_container_width=True
                )

        with tab_value:
            st.subheader("💎 Legendary Value Investing Screener")
            v_col1, v_col2 = st.columns([1, 2])
            v_col1.metric("Benjamin Graham Number", f"{curr_sym}{graham_number:.2f}")
            if graham_number > current_price: v_col2.success(f"✅ **Undervalued:** Trading below Graham's intrinsic defense value by {((graham_number-current_price)/current_price)*100:.1f}%")
            elif graham_number > 0: v_col2.warning(f"⚠️ **Overvalued:** Trading above Graham's intrinsic defense value by {((current_price-graham_number)/graham_number)*100:.1f}%")
            else: v_col2.error("❌ Invalid data for Graham calculation (Negative EPS or BVPS detected).")
            
            st.markdown("#### Dividend Discount Model (Gordon Growth)")
            d_col1, d_col2 = st.columns(2)
            d_col1.metric("Current Dividend Yield", f"{dividend_yield*100:.2f}%" if dividend_yield else "N/A")
            if dividend_rate and dividend_rate > 0:
                assumed_div_growth = 0.03
                if calculated_wacc > assumed_div_growth:
                    gordon_value = (dividend_rate * (1 + assumed_div_growth)) / (calculated_wacc - assumed_div_growth)
                    d_col2.metric(f"DDM Fair Value (Assumes {assumed_div_growth*100}% Growth)", f"{curr_sym}{gordon_value:.2f}")
                else:
                    d_col2.warning("Cost of Capital is lower than assumed dividend growth; DDM invalid.")
            else:
                d_col2.info("Asset does not pay a valid dividend for DDM analysis.")

            ebit, ev = deep_metrics.get('ebit'), info.get('enterpriseValue')
            earnings_yield = (ebit / ev * 100) if ebit and ev and ev > 0 else 0
            st.metric("Acquirer's Multiple (Earnings Yield)", f"{earnings_yield:.2f}%", help="EBIT / Enterprise Value. Metric used by private equity to find cash-cow targets.")
            
            st.markdown("#### Piotroski F-Score (Proxy Metric)")
            st.progress(f_score / 4)
            st.caption(f"Estimated Score: {f_score} / 4 (Based on Net Income, OCF, OCF > NI, and ROA signals.)")

        with tab_dcf:
            st.subheader("Valuation Architecture Deck")
            if is_financial:
                st.warning("🏛️ WACC-based DCF is structurally invalid for Financial Institutions.")
                st.markdown("**Quantitative Rationale:** Banks treat debt as raw operational material rather than capital leverage. Free Cash Flow to Firm (FCFF) calculations are distorted by operational deposit inflows.")
                st.markdown("#### Structural Equity Valuation via DDM")
                if dividend_rate and dividend_rate > 0:
                    ke_bank = current_rf + (beta_raw * global_erp)
                    g_bank = 0.04
                    if ke_bank > g_bank:
                        ddm_val = (dividend_rate * (1 + g_bank)) / (ke_bank - g_bank)
                        st.metric("Implied Dividend Fair Value", f"{curr_sym}{ddm_val:.2f}")
                    else:
                        st.info("Cost of Equity is below baseline macro growth; DDM unstable.")
                else:
                    st.info("Asset does not distribute active dividends. Utilize Excess Returns framework offline.")
            else:
                st.markdown("#### Capital Asset Pricing Model & Dynamic Cost of Capital")
                w_col1, w_col2, w_col3 = st.columns(3)
                w_col1.metric("Cost of Equity (Ke)", f"{cost_of_equity*100:.2f}%", help=f"Rf ({current_rf*100:.2f}%) + Beta ({beta_raw:.2f}) * ERP ({global_erp*100:.2f}%)")
                w_col2.metric("Effective Cost of Debt (Kd)", f"{cost_of_debt*100:.2f}%", help=f"Source: {kd_source}")
                w_col3.metric("True Blended WACC", f"{calculated_wacc*100:.2f}%", help=f"Tax Rate: {effective_tax_rate*100:.1f}% ({tax_source}) | Equity Weight: {weight_equity*100:.1f}%")
                
                st.markdown("---")
                st.markdown("#### Multi-Stage Free Cash Flow Modeler (Fade / H-Model Architecture)")
                dcf_mode = st.radio("Valuation Growth Engine Mode:", ["Multi-Stage Fade (H-Model Growth)", "Constant Growth Rate"], horizontal=True)
                
                dcf_col1, dcf_col2, dcf_col3 = st.columns(3)
                ui_fcf = dcf_col1.number_input("Base FCF Override (Millions)", value=float(fcf_base), step=10.0)
                if ui_fcf <= 0:
                    st.caption("⚠️ Negative reported FCF detected. Normalizing baseline via 12% operating revenue proxy.")
                    ui_fcf = float(raw_rev * 0.12) / 1000000.0
                
                default_wacc_ui = min(max(float(calculated_wacc*100), 5.0), 30.0)
                ui_wacc = dcf_col2.slider("Discount Rate (WACC %)", 5.0, 30.0, default_wacc_ui, 0.5) / 100
                ui_t_growth = dcf_col3.slider("Terminal Growth Rate (%)", 1.0, 8.0, 4.0, 0.5) / 100
                
                ui_dcf_results = {}
                scenario_configs = {
                    "Bear Case": {"initial_g": 0.04, "color": "#ef4444"},
                    "Base Case": {"initial_g": 0.08, "color": "#3b82f6"},
                    "Bull Case": {"initial_g": 0.14, "color": "#10b981"}
                }
                
                for sc, conf in scenario_configs.items():
                    init_g = conf["initial_g"]
                    if dcf_mode == "Multi-Stage Fade (H-Model Growth)":
                        # Years 1-2: High growth; Years 3-5: Linear fade toward terminal rate
                        g_schedule = [
                            init_g, 
                            init_g, 
                            init_g - (init_g - ui_t_growth) * 0.33, 
                            init_g - (init_g - ui_t_growth) * 0.66, 
                            ui_t_growth + 0.005
                        ]
                    else:
                        g_schedule = [init_g] * 5
                    
                    # Project compounding FCFs along schedule
                    cfs = []
                    curr_cf = ui_fcf
                    for yr_g in g_schedule:
                        curr_cf *= (1 + yr_g)
                        cfs.append(curr_cf)
                        
                    pv = sum([cfs[t] / ((1 + ui_wacc) ** (t + 1)) for t in range(5)])
                    tv = (cfs[-1] * (1 + ui_t_growth)) / max(ui_wacc - ui_t_growth, 0.015)
                    ui_dcf_results[sc] = (pv + (tv / ((1 + ui_wacc) ** 5))) / shares_for_calc

                st.plotly_chart(px.bar(pd.DataFrame(list(ui_dcf_results.items()), columns=["Scenario", "Target Price"]), 
                                       x="Scenario", y="Target Price", text_auto=".2f", color="Scenario", 
                                       color_discrete_map={"Bear Case": "#ef4444", "Base Case": "#3b82f6", "Bull Case": "#10b981"}), use_container_width=True)
                
                st.markdown("#### Reverse DCF (Market Implied Growth Rate)")
                implied_g_range = np.linspace(-0.10, 0.30, 400)
                closest_diff, implied_g_ans = float('inf'), 0
                for test_g in implied_g_range:
                    if ui_wacc > test_g:
                        cfs = [ui_fcf * ((1 + test_g) ** y) for y in range(1, 6)]
                        pv = sum([cfs[t] / ((1 + ui_wacc) ** (t + 1)) for t in range(5)])
                        tv = (cfs[-1] * (1 + test_g)) / max(ui_wacc - test_g, 0.015)
                        test_price = (pv + (tv / ((1 + ui_wacc) ** 5))) / shares_for_calc
                        if abs(test_price - current_price) < closest_diff:
                            closest_diff = abs(test_price - current_price)
                            implied_g_ans = test_g
                st.metric("Market Implied Growth Rate", f"{implied_g_ans*100:.2f}%", help="If the company grows slower than this, the stock is currently overvalued.")
                
                st.markdown("#### WACC vs. Terminal Growth Sensitivity Matrix")
                wacc_range = np.arange(max(0.05, ui_wacc - 0.02), ui_wacc + 0.03, 0.01)
                tg_range = np.arange(max(0.01, ui_t_growth - 0.015), ui_t_growth + 0.02, 0.005)
                matrix = np.zeros((len(wacc_range), len(tg_range)))
                for i, w in enumerate(wacc_range):
                    for j, t in enumerate(tg_range):
                        cfs = [ui_fcf * ((1 + 0.08) ** y) for y in range(1, 6)]
                        pv = sum([cfs[year] / ((1 + w) ** (year + 1)) for year in range(5)])
                        tv = (cfs[-1] * (1 + t)) / max(w - t, 0.015)
                        matrix[i, j] = (pv + (tv / ((1 + w) ** 5))) / shares_for_calc
                fig_heat = px.imshow(matrix, labels=dict(x="Terminal Growth Rate", y="Discount Rate (WACC)", color="Implied Price"), 
                                     x=[f"{x*100:.1f}%" for x in tg_range], y=[f"{y*100:.1f}%" for y in wacc_range], 
                                     text_auto=".2f", color_continuous_scale="RdYlGn")
                st.plotly_chart(fig_heat, use_container_width=True)

        with tab_health:
            st.subheader("🏛️ Corporate Health & Forensics")
            st.markdown("#### Value Creation Engine (ROIC vs. Cost of Capital)")
            st.caption(f"**Invested Capital Architecture:** {ic_source} | Effective Tax Rate: {effective_tax_rate*100:.1f}%")
            
            if ebit_val > 0 and invested_capital > 0:
                nopat = ebit_val * (1 - effective_tax_rate)
                roic = nopat / invested_capital
                roic_wacc_spread = roic - calculated_wacc
                
                r_col1, r_col2 = st.columns([1, 2])
                r_col1.metric("Return on Invested Capital (ROIC)", f"{roic*100:.2f}%")
                if roic_wacc_spread > 0: 
                    r_col2.success(f"✅ **Value Creator:** ROIC exceeds WACC by {roic_wacc_spread*100:.2f}%. True economic value created.")
                else: 
                    r_col2.error(f"🚨 **Value Destroyer:** ROIC is below WACC by {abs(roic_wacc_spread)*100:.2f}%. Cost of capital exceeds return.")

            st.markdown("#### Altman Z-Score (Bankruptcy Probability Model)")
            if is_financial:
                st.info("Altman Z-Score analysis bypassed. Model metrics are incompatible with banking leverage profiles.")
            elif z_score is not None:
                z_col1, z_col2 = st.columns([1, 2])
                z_col1.metric(f"{model_type}", f"{z_score:.2f}")
                if z_score >= z_safe_limit: 
                    z_col2.success("✅ **Safe Zone:** Structural insolvency risk is mathematically remote.")
                elif z_distress_limit <= z_score < z_safe_limit: 
                    z_col2.warning("⚠️ **Grey Zone:** Structural friction detected. Monitor capitalization trends closely.")
                else: 
                    z_col2.error("🚨 **Distress Zone:** High profile vulnerability. Restructuring indicators present.")
            else:
                st.info("Insufficient deep balance sheet data to calculate Altman Z-Score.")
            
            st.markdown("---")
            if dupont_data['valid']:
                st.markdown("#### 3-Stage DuPont ROE Decomposition")
                dp1, dp2, dp3, dp4 = st.columns(4)
                dp1.metric("Net Profit Margin", f"{dupont_data['npm']:.2f}%", help="Operational Margin Efficiency")
                dp2.metric("Asset Turnover", f"{dupont_data['ato']:.2f}x", help="Asset Utilization Efficiency")
                dp3.metric("Equity Multiplier", f"{dupont_data['em']:.2f}x", help="Financial Leverage Multiplier")
                dp4.metric("Deconstructed ROE", f"{dupont_data['roe']:.2f}%", help="NPM * ATO * EM")

        with tab_mpt:
            st.subheader("Modern Portfolio Theory (MPT) Optimization")
            if len(all_tickers_to_compare) >= 2:
                with st.spinner("Running SLSQP Constrained Optimization over Peer Cohort..."):
                    mpt_data = pull_peer_action(all_tickers_to_compare, time_horizon)
                    if not mpt_data.empty and isinstance(mpt_data, pd.DataFrame):
                        mpt_data = mpt_data.ffill().bfill()
                        weekly_data = mpt_data.resample('W').last()
                        ret = weekly_data.pct_change().dropna()
                        valid_tickers = list(ret.columns)
                        num_assets = len(valid_tickers)
                        
                        if num_assets >= 2 and len(ret) > 10:
                            st.markdown("#### Inter-Asset Correlation Matrix (Weekly Resampled)")
                            fig_corr = px.imshow(ret.corr(), text_auto=".2f", color_continuous_scale="Blues", aspect="auto")
                            st.plotly_chart(fig_corr, use_container_width=True)
                            
                            ann_ret = ret.mean().values * 52
                            cov_matrix = ret.cov().values * 52
                            
                            def portfolio_performance(weights, mean_returns, cov_mat):
                                returns = np.sum(mean_returns * weights)
                                std_dev = np.sqrt(np.dot(weights.T, np.dot(cov_mat, weights)))
                                return returns, std_dev
                                
                            def negative_sharpe(weights, mean_returns, cov_mat, rf_rate):
                                p_ret, p_std = portfolio_performance(weights, mean_returns, cov_mat)
                                return -(p_ret - rf_rate) / p_std if p_std > 0 else 0
                                
                            constraints = ({'type': 'eq', 'fun': lambda x: np.sum(x) - 1})
                            bounds = tuple((0.0, 1.0) for _ in range(num_assets))
                            init_guess = [1.0 / num_assets] * num_assets
                            
                            opt_res = sco.minimize(negative_sharpe, init_guess, args=(ann_ret, cov_matrix, current_rf), 
                                                   method='SLSQP', bounds=bounds, constraints=constraints)
                            opt_weights = opt_res.x
                            opt_ret, opt_std = portfolio_performance(opt_weights, ann_ret, cov_matrix)
                            max_sharpe = (opt_ret - current_rf) / opt_std
                            
                            m1, m2, m3 = st.columns(3)
                            m1.metric("Optimized Expected Return", f"{opt_ret * 100:.2f}%")
                            m2.metric("Optimized Annual Risk", f"{opt_std * 100:.2f}%")
                            m3.metric("Maximized Sharpe Ratio", f"{max_sharpe:.2f}")
                            
                            res = np.zeros((3, 2000))
                            for i in range(2000):
                                w = np.random.random(num_assets); w /= np.sum(w)
                                p_ret = np.sum(ann_ret * w); p_std = np.sqrt(np.dot(w.T, np.dot(cov_matrix, w)))
                                res[0,i], res[1,i] = p_ret, p_std
                                res[2,i] = (p_ret - current_rf) / p_std 
                                
                            st.markdown("#### Efficient Frontier (SLSQP Global Maximum)")
                            fig_mpt = px.scatter(x=res[1,:], y=res[0,:], color=res[2,:], 
                                                 labels={'x': 'Risk', 'y': 'Return', 'color': 'Sharpe'}, title="Efficient Frontier Simulation")
                            fig_mpt.add_trace(go.Scatter(x=[opt_std], y=[opt_ret], mode='markers', marker=dict(color='red', size=18, symbol='star'), name='SLSQP Optimal'))
                            fig_mpt.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
                            st.plotly_chart(fig_mpt, use_container_width=True)
                            
                            st.markdown(f"#### SLSQP Optimal Capital Deployment (Assuming {curr_sym}{portfolio_capital:,.2f})")
                            alloc_data = []
                            for idx, ticker in enumerate(valid_tickers):
                                alloc_data.append({"Asset": ticker, "Weight": opt_weights[idx], "Capital": opt_weights[idx] * portfolio_capital})
                            
                            fig_pie = px.pie(pd.DataFrame(alloc_data), names="Asset", values="Weight", hole=0.4, title="Target Capital Allocation Weights")
                            fig_pie.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
                            st.plotly_chart(fig_pie, use_container_width=True)
                            st.dataframe(pd.DataFrame(alloc_data).style.format({"Weight": "{:.2%}", "Capital": f"{curr_sym}"+"{:,.2f}"}), use_container_width=True)
                        else:
                            st.warning("Insufficient overlapping historical data across the peer group.")
            else: st.warning("Add peers to run Portfolio Optimization.")

        with tab_ml:
            st.subheader(f"🤖 Machine Learning Price Forecast ({selected_ticker})")
            if not df_market.empty:
                returns = df_market['close_price'].pct_change().dropna()
                mu, sigma = returns.mean(), returns.std()
                days_to_predict, simulations = 30, 100
                last_price = df_market['close_price'].iloc[-1]
                
                sim_df = np.zeros((days_to_predict, simulations))
                sim_df[0] = last_price
                for t in range(1, days_to_predict):
                    shock = np.random.normal(loc=0, scale=1, size=simulations)
                    sim_df[t] = sim_df[t-1] * np.exp((mu - (sigma**2) / 2) + sigma * shock)
                
                fig_ml = go.Figure()
                for i in range(simulations):
                    fig_ml.add_trace(go.Scatter(x=np.arange(days_to_predict), y=sim_df[:, i], mode='lines', line=dict(color='#3b82f6', width=1), opacity=0.1, showlegend=False))
                
                mean_path = sim_df.mean(axis=1)
                upper_bound = np.percentile(sim_df, 95, axis=1)
                lower_bound = np.percentile(sim_df, 5, axis=1)
                
                fig_ml.add_trace(go.Scatter(x=np.arange(days_to_predict), y=mean_path, mode='lines', line=dict(color='#ef4444', width=3), name="Expected Mean Path"))
                fig_ml.add_trace(go.Scatter(x=np.arange(days_to_predict), y=upper_bound, mode='lines', line=dict(color='#10b981', width=2, dash='dash'), name="95th Percentile Bounds"))
                fig_ml.add_trace(go.Scatter(x=np.arange(days_to_predict), y=lower_bound, mode='lines', line=dict(color='#ef4444', width=2, dash='dash'), showlegend=False))
                
                fig_ml.update_layout(title=f"30-Day ML Monte Carlo Forecast Envelope ({selected_ticker})", plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', xaxis_title="Days into Future", yaxis_title=f"Projected Price ({currency})")
                st.plotly_chart(fig_ml, use_container_width=True)

        with tab_bs:
            st.subheader(f"Institutional Options Desk ({selected_ticker})")
            st.markdown("#### Live Volatility Skew & Put/Call Ratio (PCR)")
            try:
                expirations = raw_ticker.options
                if expirations:
                    chain = raw_ticker.option_chain(expirations[0])
                    live_calls, live_puts = chain.calls, chain.puts
                    total_call_vol = live_calls['volume'].sum() if 'volume' in live_calls.columns else 1
                    total_put_vol = live_puts['volume'].sum() if 'volume' in live_puts.columns else 0
                    pcr = total_put_vol / total_call_vol if total_call_vol > 0 else 0
                    
                    pcr_col1, pcr_col2 = st.columns(2)
                    pcr_col1.metric(f"Live Put/Call Ratio (Exp: {expirations[0]})", f"{pcr:.2f}")
                    if pcr > 1: pcr_col2.error("🚨 **Bearish Sentiment:** More Puts are being traded than Calls.")
                    elif pcr < 0.7: pcr_col2.success("✅ **Bullish Sentiment:** Heavy Call volume relative to Puts.")
                    else: pcr_col2.info("⚖️ **Neutral Sentiment:** Option volume flow is balanced.")
                    
                    fig_skew = px.scatter(live_calls, x='strike', y='impliedVolatility', size='openInterest', color='impliedVolatility', 
                                          title=f"Implied Volatility Smile for Nearest Expiry: {expirations[0]}", color_continuous_scale="Viridis")
                    fig_skew.add_vline(x=current_price, line_dash="dash", line_color="gray", annotation_text="Current Price")
                    fig_skew.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', xaxis_title="Strike Price", yaxis_title="Implied Volatility")
                    st.plotly_chart(fig_skew, use_container_width=True)
                else: st.info("No live options chains available for this ticker.")
            except: st.warning("Could not fetch live options data from exchange.")

            st.markdown("#### Theoretical Premium Modeler")
            bs_col1, bs_col2, bs_col3, bs_col4 = st.columns(4)
            K = bs_col1.number_input("Strike Price (K)", value=float(current_price * 1.05), step=1.0)
            T = bs_col2.slider("Time to Expiry (Years)", 0.01, 5.0, 1.0, 0.05)
            r = bs_col3.slider("Risk-Free Rate (%)", 1.0, 10.0, float(current_rf*100), 0.1) / 100
            sigma = bs_col4.slider("Implied Volatility (%)", 5.0, 150.0, 30.0, 1.0) / 100
            
            d1 = (np.log(current_price / K) + (r + 0.5 * sigma ** 2) * T) / (sigma * np.sqrt(T))
            d2 = d1 - sigma * np.sqrt(T)
            call_price = current_price * si.norm.cdf(d1, 0.0, 1.0) - K * np.exp(-r * T) * si.norm.cdf(d2, 0.0, 1.0)
            put_price = K * np.exp(-r * T) * si.norm.cdf(-d2, 0.0, 1.0) - current_price * si.norm.cdf(-d1, 0.0, 1.0)
            call_delta = si.norm.cdf(d1, 0.0, 1.0)
            gamma = si.norm.pdf(d1, 0.0, 1.0) / (current_price * sigma * np.sqrt(T))
            vega = current_price * si.norm.pdf(d1, 0.0, 1.0) * np.sqrt(T) / 100
            call_theta = (-current_price * si.norm.pdf(d1, 0.0, 1.0) * sigma / (2 * np.sqrt(T)) - r * K * np.exp(-r * T) * si.norm.cdf(d2, 0.0, 1.0)) / 365
            
            call_col, put_col = st.columns(2)
            call_col.metric("Theoretical Call Premium", f"{curr_sym}{call_price:.2f}")
            put_col.metric("Theoretical Put Premium", f"{curr_sym}{put_price:.2f}")
            
            g_col1, g_col2, g_col3, g_col4 = st.columns(4)
            g_col1.metric("Delta (Call)", f"{call_delta:.3f}")
            g_col2.metric("Gamma", f"{gamma:.4f}")
            g_col3.metric("Theta (Daily Decay)", f"{curr_sym}{call_theta:.3f}")
            g_col4.metric("Vega", f"{curr_sym}{vega:.3f}")

        with tab_tech:
            st.subheader(f"Quantitative Trading Desk: Microstructure & Algos ({selected_ticker})")
            if not df_market.empty:
                st.markdown("#### Volume Profile (VPVR) - Institutional Price Nodes")
                tech_df = df_market.copy().sort_values('date')
                fig_vp = px.histogram(tech_df, y='close_price', x='volume', orientation='h', nbins=30, color_discrete_sequence=['#94a3b8'], title=f"Volume Accumulation by Price Level ({selected_ticker})")
                fig_vp.add_hline(y=current_price, line_dash="solid", line_color="#3b82f6", annotation_text="Current Price")
                fig_vp.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', yaxis_title=f"Price Level ({currency})", xaxis_title="Total Historical Volume Traded")
                st.plotly_chart(fig_vp, use_container_width=True)

                st.markdown("#### Momentum Oscillators")
                tech_col1, tech_col2, tech_col3 = st.columns(3)
                bb_window = tech_col1.slider("Bollinger Window", 10, 50, 20, 1)
                rsi_window = tech_col2.slider("RSI Lookback", 7, 30, 14, 1)
                macd_fast = tech_col3.slider("MACD Fast EMA", 5, 20, 12, 1)

                tech_df['BB_Middle'] = tech_df['close_price'].rolling(window=bb_window).mean()
                tech_df['BB_Std'] = tech_df['close_price'].rolling(window=bb_window).std()
                tech_df['BB_Upper'] = tech_df['BB_Middle'] + (2.0 * tech_df['BB_Std'])
                tech_df['BB_Lower'] = tech_df['BB_Middle'] - (2.0 * tech_df['BB_Std'])
                
                delta = tech_df['close_price'].diff()
                gain = (delta.where(delta > 0, 0)).rolling(window=rsi_window).mean()
                loss = (-delta.where(delta < 0, 0)).rolling(window=rsi_window).mean()
                rs = gain / (loss + 1e-10) 
                tech_df['RSI'] = 100 - (100 / (1 + rs))
                
                ema_fast = tech_df['close_price'].ewm(span=macd_fast, adjust=False).mean()
                ema_slow = tech_df['close_price'].ewm(span=26, adjust=False).mean()
                tech_df['MACD'] = ema_fast - ema_slow
                tech_df['Signal_Line'] = tech_df['MACD'].ewm(span=9, adjust=False).mean()
                tech_df['MACD_Histogram'] = tech_df['MACD'] - tech_df['Signal_Line']
                tech_df = tech_df.dropna()
                
                fig_bb = go.Figure()
                fig_bb.add_trace(go.Scatter(x=tech_df['date'], y=tech_df['close_price'], name='Close Price', line=dict(color='#3b82f6', width=2)))
                fig_bb.add_trace(go.Scatter(x=tech_df['date'], y=tech_df['BB_Upper'], name='Upper Band', line=dict(color='#ef4444', width=1, dash='dot')))
                fig_bb.add_trace(go.Scatter(x=tech_df['date'], y=tech_df['BB_Lower'], name='Lower Band', line=dict(color='#10b981', width=1, dash='dot')))
                fig_bb.update_layout(title=f"Volatility Channels ({selected_ticker})", plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
                st.plotly_chart(fig_bb, use_container_width=True)

        with tab_season:
            st.subheader(f"📅 Quantitative Seasonality & Regime Tracking ({selected_ticker})")
            if not df_market.empty:
                season_df = df_market.copy()
                season_df['Daily_Ret'] = season_df['close_price'].pct_change()
                season_df['Month'] = season_df['date'].dt.month_name().str[:3]
                season_df['Year'] = season_df['date'].dt.year
                season_df = season_df.dropna()
                
                pivot_season = pd.pivot_table(season_df, values='Daily_Ret', index='Year', columns='Month', aggfunc=lambda x: (np.prod(1+x)-1)*100)
                months_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                pivot_season = pivot_season.reindex(columns=[m for m in months_order if m in pivot_season.columns])
                
                fig_season = px.imshow(pivot_season, text_auto=".1f", color_continuous_scale="RdYlGn", title=f"Historical Monthly Return Heatmap (%)")
                fig_season.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
                st.plotly_chart(fig_season, use_container_width=True)

        with tab_risk:
            st.subheader(f"Institutional Risk Engine: Value at Risk (VaR) & Expected Shortfall ({selected_ticker})")
            if not df_market.empty:
                risk_df = df_market.copy().sort_values('date')
                risk_df['Daily Return'] = risk_df['close_price'].pct_change()
                risk_df = risk_df.dropna()
                
                var_95, var_99 = np.percentile(risk_df['Daily Return'], 5), np.percentile(risk_df['Daily Return'], 1)
                cvar_95 = risk_df[risk_df['Daily Return'] <= var_95]['Daily Return'].mean()
                
                risk_df['Cumulative Max'] = risk_df['close_price'].cummax()
                risk_df['Drawdown'] = (risk_df['close_price'] - risk_df['Cumulative Max']) / risk_df['Cumulative Max']
                max_drawdown = risk_df['Drawdown'].min()
                
                r_col1, r_col2, r_col3 = st.columns(3)
                r_col1.metric("VaR (95%) Threshold", f"{var_95 * 100:.2f}%")
                r_col2.metric("Conditional VaR (Expected Shortfall)", f"{cvar_95 * 100:.2f}%")
                r_col3.metric("Max Historical Drawdown", f"{max_drawdown * 100:.2f}%")
                
                fig_dd = px.area(risk_df, x='date', y='Drawdown', title="Underwater Curve (Drawdown Profile)", color_discrete_sequence=['#ef4444'])
                fig_dd.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', yaxis_tickformat='.2%')
                st.plotly_chart(fig_dd, use_container_width=True)

        # ------------------------------------------------------------------
        # TAB 12: STATISTICAL ARBITRAGE WITH COINTEGRATION & ADF TESTING
        # ------------------------------------------------------------------
        with tab_arb:
            st.subheader("🎭 Statistical Arbitrage & Cointegration Desk")
            default_peer = resolve_automatic_peer(selected_ticker, info.get('sector', ''), info.get('industry', ''))
            arb_ui_col1, arb_ui_col2 = st.columns([1, 2])
            with arb_ui_col1:
                arb_pair = st.text_input("📊 Target Arbitrage Counter-Pair Ticker:", value=default_peer, 
                                         help="Automatically mapped based on sector symmetry. Override manually as needed.")
            
            if arb_pair:
                with st.spinner(f"Pulling dynamic arbitrage spread and cointegration analytics for {arb_pair}..."):
                    try:
                        arb_data = pull_peer_action([arb_pair.strip()], time_horizon)
                        if not arb_data.empty:
                            arb_df = pd.DataFrame(arb_data).dropna().reset_index()
                            arb_df.columns = ['date', 'arb_close']
                            arb_df['date'] = pd.to_datetime(arb_df['date']).dt.tz_localize(None)
                            
                            pair_df = pd.merge(df_market[['date', 'close_price']], arb_df, on='date', how='inner')
                            if not pair_df.empty and len(pair_df) > 20:
                                p1 = pair_df['close_price'].values
                                p2 = pair_df['arb_close'].values
                                
                                # 1. OLS Hedge Ratio Beta & Spread Residuals
                                cov_12 = np.cov(p1, p2)[0, 1]
                                var_2 = np.var(p2)
                                hedge_ratio = cov_12 / var_2 if var_2 > 0 else 1.0
                                alpha = np.mean(p1) - hedge_ratio * np.mean(p2)
                                spread_residuals = p1 - (alpha + hedge_ratio * p2)
                                pair_df['Hedged_Spread'] = spread_residuals
                                
                                # 2. Augmented Dickey-Fuller Cointegration Test
                                t_stat, p_val, is_coint = run_adf_stationarity_test(spread_residuals)
                                
                                # Display Cointegration Diagnostics
                                c1, c2, c3 = st.columns(3)
                                c1.metric("Optimal OLS Hedge Ratio (β)", f"{hedge_ratio:.3f}", help="Hedge Ratio: Buy 1 share of Target, Short β shares of Pair.")
                                c2.metric("ADF Test Statistic (t)", f"{t_stat:.2f}", help="MacKinnon 5% Critical Value is -2.86. Value must be more negative to confirm mean-reversion.")
                                if is_coint:
                                    c3.success(f"✅ **Cointegrated** (p = {p_val:.3f}) — Stationary Mean-Reverting Spread.")
                                else:
                                    c3.warning(f"⚠️ **Non-Cointegrated** (p = {p_val:.3f}) — Spurious Spread / Random Walk.")
                                
                                # 3. Standardized Rolling Z-Score on Hedged Residual Spread
                                roll_mean = pair_df['Hedged_Spread'].rolling(window=20).mean()
                                roll_std = pair_df['Hedged_Spread'].rolling(window=20).std()
                                pair_df['Z_Score'] = (pair_df['Hedged_Spread'] - roll_mean) / (roll_std + 1e-8)
                                pair_df = pair_df.dropna()
                                
                                fig_arb = px.line(pair_df, x='date', y='Z_Score', 
                                                  title=f"Cointegrated Spread Z-Score: {selected_ticker} vs. {arb_pair.strip()} (β={hedge_ratio:.3f})", 
                                                  color_discrete_sequence=['#a855f7'])
                                fig_arb.add_hline(y=2.0, line_dash="dash", line_color="#ef4444", annotation_text="Sell Target / Buy Pair Spread Threshold (+2.0σ)")
                                fig_arb.add_hline(y=-2.0, line_dash="dash", line_color="#10b981", annotation_text="Buy Target / Sell Pair Spread Threshold (-2.0σ)")
                                fig_arb.add_hline(y=0.0, line_dash="dot", line_color="#94a3b8", annotation_text="Mean Reversion Equilibrium")
                                fig_arb.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)')
                                st.plotly_chart(fig_arb, use_container_width=True)
                            else:
                                st.warning("Insufficient overlapping trading history to establish cointegration.")
                        else:
                            st.warning("Selected pair returned empty historical dataframe.")
                    except Exception as e: 
                        st.error(f"Execution failed for specified pair asset: {str(e)}")

        with tab_insider:
            st.subheader("🕵️ Smart Money Tracker (Institutional & Insider Activity)")
            has_insider_data = False
            try:
                holders = raw_ticker.institutional_holders
                if holders is not None and not holders.empty:
                    st.markdown("#### Top Institutional Holders")
                    st.dataframe(holders, use_container_width=True)
                    has_insider_data = True
            except: pass
            
            try:
                insider = raw_ticker.insider_purchases
                if insider is not None and not insider.empty:
                    st.markdown("#### Recent Insider Transaction Summary")
                    st.dataframe(insider, use_container_width=True)
                    has_insider_data = True
            except: pass
            
            if not has_insider_data:
                st.markdown("<br>", unsafe_allow_html=True)
                st.info("⚠️ SEC/SEBI Filing Data Currently Unavailable via Open-Source Feeds.")

        with tab_macro:
            st.subheader("🌐 Global Macroeconomic Regime")
            with st.spinner("Fetching Treasury Yields and Volatility Index..."):
                try:
                    yield_data = pull_macro_regime(time_horizon)
                    if not yield_data.empty:
                        yield_data = yield_data.dropna()
                        yield_data['Yield_Spread'] = yield_data['^TNX'] - yield_data['^IRX']
                        
                        m_col1, m_col2 = st.columns(2)
                        with m_col1:
                            fig_spread = px.area(x=yield_data.index, y=yield_data['Yield_Spread'], title="10Y-3M Yield Curve Spread", color_discrete_sequence=['#a855f7'])
                            fig_spread.add_hline(y=0, line_dash="solid", line_color="#ef4444", annotation_text="Inversion Line")
                            fig_spread.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', yaxis_title="Spread Basis Points")
                            st.plotly_chart(fig_spread, use_container_width=True)
                        with m_col2:
                            fig_vix = px.line(x=yield_data.index, y=yield_data['^VIX'], title="CBOE Volatility Index (VIX)", color_discrete_sequence=['#ef4444'])
                            fig_vix.add_hline(y=20, line_dash="dash", line_color="#94a3b8", annotation_text="High Fear Threshold")
                            fig_vix.update_layout(plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', yaxis_title="VIX Level")
                            st.plotly_chart(fig_vix, use_container_width=True)
                except: st.warning("Macro data currently unavailable.")

        with tab_ai:
            st.subheader(f"🧠 Institutional AI Synthesis ({selected_ticker})")
            st.markdown("Synthesize deterministic quantitative models into a structured narrative.")

            def _render_thesis_display(thesis_data):
                col1, col2 = st.columns([1, 3])
                with col1:
                    v_color = "green" if thesis_data.get('verdict') == "BUY" else ("red" if thesis_data.get('verdict') == "SELL" else "orange")
                    st.markdown(f"### Verdict: :{v_color}[{thesis_data.get('verdict', 'N/A')}]")
                with col2:
                    st.markdown(f"**Rationale:** *{thesis_data.get('target_rationale', 'N/A')}*")

                st.divider()

                left_flow, right_flow = st.columns(2)
                with left_flow:
                    st.markdown("### 📊 Valuation & Variant Perception")
                    st.write(thesis_data.get('valuation_case', 'N/A'))
                    st.info(f"**Variant Perception:** {thesis_data.get('variant_perception', 'N/A')}")

                with right_flow:
                    st.markdown("### ⚠️ Risk Assessment & Data Grounding")
                    st.error(f"**Primary Structural Risk:** {thesis_data.get('core_risk_factor', 'N/A')}")
                    st.markdown("**Core Data Pillars Used:**")
                    for item in thesis_data.get('quantitative_grounding', []):
                        st.markdown(f"- `{item}`")

            if st.button("Run Quantitative Synthesis", use_container_width=True):
                v_price = f"{curr_sym}{current_price:.2f}" if current_price is not None else "N/A"
                v_dcf = f"{curr_sym}{ui_dcf_results.get('Base Case', 0):.2f}" if 'ui_dcf_results' in locals() and ui_dcf_results else "N/A"
                v_wacc = f"{calculated_wacc*100:.2f}%" if calculated_wacc is not None else "N/A"
                v_zscore = f"{z_score:.2f} ({model_type})" if z_score is not None else "N/A"
                v_fscore = f"{f_score}/4" if f_score is not None else "N/A"
                v_roic = f"{roic_wacc_spread*100:.2f}%" if roic_wacc_spread is not None else "N/A"
                v_rsi = f"{tech_df['RSI'].iloc[-1]:.2f}" if 'tech_df' in locals() and not tech_df.empty and 'RSI' in tech_df.columns else "N/A"

                metrics_summary = (
                    f"- Current Price: {v_price}\n"
                    f"- DCF Base Case Implied Value: {v_dcf}\n"
                    f"- Cost of Capital / WACC: {v_wacc}\n"
                    f"- Altman Z-Score: {v_zscore}\n"
                    f"- Piotroski F-Score: {v_fscore}\n"
                    f"- ROIC vs WACC Spread: {v_roic}\n"
                    f"- RSI (14-Day): {v_rsi}\n"
                )
                
                with st.spinner("Executing multi-variable semantic analysis via Gemini..."):
                    try:
                        thesis_data = generate_ai_thesis(selected_ticker, full_name, metrics_summary)
                    except RuntimeError as e:
                        thesis_data = f"ERROR: {e}"

                    if isinstance(thesis_data, dict):
                        st.session_state.ai_thesis_store[selected_ticker] = thesis_data
                    elif isinstance(thesis_data, str) and thesis_data.startswith("ERROR:"):
                        st.error(f"API Diagnostics: {thesis_data}")
                    else:
                        st.warning("Engine failed to parse structured thesis. Please try again.")

            _persisted_thesis = st.session_state.get("ai_thesis_store", {}).get(selected_ticker)
            if _persisted_thesis:
                st.caption(f"✅ Cached synthesis for {selected_ticker} — included in the PDF Tear Sheet above.")
                _render_thesis_display(_persisted_thesis)

        # PDF Tear Sheet render deferred block
        cached_thesis = st.session_state.ai_thesis_store.get(selected_ticker)
        pdf_ctx = {
            "ticker": selected_ticker, "full_name": full_name, "sector": info.get('sector', 'N/A'),
            "industry": info.get('industry', 'N/A'), "exchange": info.get('exchange', 'N/A'),
            "curr_sym": curr_sym, "currency": currency, "current_price": current_price,
            "market_cap": market_cap, "beta": beta_raw, "pe": info.get('trailingPE'),
            "high_52": info.get('fiftyTwoWeekHigh'), "low_52": info.get('fiftyTwoWeekLow'),
            "div_yield": dividend_yield, "dcf_results": pdf_dcf, "graham_number": graham_number,
            "dupont": dupont_data, "z_score": z_score, "z_model_type": model_type,
            "z_safe_limit": z_safe_limit, "z_distress_limit": z_distress_limit,
            "f_score": f_score, "roic": roic, "wacc": calculated_wacc, "df_market": df_market,
            "ai_thesis": cached_thesis,
        }
        pdf_bytes = generate_pdf_report(pdf_ctx)
        with pdf_slot:
            st.download_button("📥 PDF Tear Sheet", data=pdf_bytes, file_name=f"{selected_ticker}_Tear_Sheet.pdf", mime="application/pdf", use_container_width=True)
            if cached_thesis:
                st.caption("Includes AI synthesis from this session.")
            else:
                st.caption("Run AI Synthesis (bottom tab) to include it in this PDF.")

